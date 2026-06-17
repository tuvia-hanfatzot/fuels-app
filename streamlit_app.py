import streamlit as st
from io import BytesIO
from copy import copy
import re
import base64
import os
from difflib import SequenceMatcher, get_close_matches

import pandas as pd

from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
from openpyxl.utils.cell import range_boundaries

from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.styles import PatternFill, Font

from datetime import datetime, date, timedelta


# ============================================================
# Streamlit config + UI styling
# ============================================================
st.set_page_config(page_title="Total Distribution → Calculations", layout="centered")

st.markdown(
    """
<style>
    .block-container { padding-top: 2rem; max-width: 900px; }
    .title { font-size: 2rem; font-weight: 800; margin-bottom: .25rem; }
    .subtitle { color: #6b7280; margin-bottom: 1.25rem; }
    .card {
        border: 1px solid rgba(0,0,0,.08);
        border-radius: 16px;
        padding: 16px 18px;
        background: rgba(255,255,255,.6);
    }
    .small { font-size: .9rem; color: #6b7280; }
</style>
""",
    unsafe_allow_html=True,
)

st.markdown('<div class="title">Total Distribution Combiner + Fuels Cleaner</div>', unsafe_allow_html=True)
st.markdown(
    '<div class="subtitle">Upload multiple Total Distribution files → combine first → then run your calculations → download final.</div>',
    unsafe_allow_html=True,
)


# ============================================================
# Settings
# ============================================================
TARGET_SHEETS = {
    "UNOPS": "UNOPS Total Distribution",
    "WFP": "WFP Total Distribution",
}

HEADER_ROWS = 2
DATA_START_ROW = 3

# If you still only want to freeze specific columns in source sheets you can keep this,
# but we now copy VALUES for *all* cells into the master anyway.
FREEZE_COLS = (3, 4)  # C, D  (kept for compatibility)

# After combining, delete columns F–I (6..9) before calculations
DELETE_COMBINED_COLS_F_TO_I_BEFORE_CALC = True


# ============================================================
# Small UI helper (optional loader video)
# ============================================================
def show_small_loader_video(placeholder, path: str, width_px: int = 220):
    try:
        with open(path, "rb") as f:
            b64 = base64.b64encode(f.read()).decode("utf-8")
        placeholder.markdown(
            f"""
            <div style="display:flex;justify-content:center;margin:12px 0;">
              <video width="{width_px}" autoplay loop muted playsinline>
                <source src="data:video/mp4;base64,{b64}" type="video/mp4">
              </video>
            </div>
            """,
            unsafe_allow_html=True,
        )
    except Exception:
        # If video missing, just ignore (no crash)
        placeholder.empty()


# ============================================================
# COMBINER HELPERS
# ============================================================
def _norm(s: str) -> str:
    s = (s or "").strip().lower()
    s = re.sub(r"\s+", " ", s)
    return s


def _find_target_sheet(wb):
    sheetnames = list(wb.sheetnames)
    norm_map = {name: _norm(name) for name in sheetnames}

    targets = {
        "UNOPS": _norm("UNOPS Total Distribution"),
        "WFP": _norm("WFP Total Distribution"),
    }

    # 1) Normalised exact match
    for kind, target_norm in targets.items():
        for real_name, real_norm in norm_map.items():
            if real_norm == target_norm:
                return kind, real_name

    # 2) Broader match
    for real_name, real_norm in norm_map.items():
        if "total" in real_norm and "distribution" in real_norm:
            if "unops" in real_norm:
                return "UNOPS", real_name
            if "wfp" in real_norm:
                return "WFP", real_name

    return None, None


def _row_has_data_row(ws, r, max_col):
    for c in range(1, max_col + 1):
        if ws.cell(r, c).value not in (None, ""):
            return True
    return False


def _first_data_row(ws, start=DATA_START_ROW, max_scan=200):
    end = min(ws.max_row, start + max_scan)
    for r in range(start, end + 1):
        if _row_has_data_row(ws, r, ws.max_column):
            return r
    return start


def _last_data_row(ws, start_row, max_col):
    for r in range(ws.max_row, start_row - 1, -1):
        if _row_has_data_row(ws, r, max_col):
            return r
    return start_row - 1


def _delete_last_text_row_in_sheet(ws, start_row, end_row, max_col):
    """
    Delete the last row that contains any value (end_row already points to last data row).
    Returns new end_row.
    """
    if end_row < start_row:
        return end_row
    ws.delete_rows(end_row, 1)
    return end_row - 1


def _copy_dimensions(src_ws, dst_ws, max_col, header_rows):
    # Column widths
    for c in range(1, max_col + 1):
        L = get_column_letter(c)
        if L in src_ws.column_dimensions:
            w = src_ws.column_dimensions[L].width
            if w is not None:
                dst_ws.column_dimensions[L].width = w

    # Header row heights
    for r in range(1, header_rows + 1):
        if r in src_ws.row_dimensions and src_ws.row_dimensions[r].height is not None:
            dst_ws.row_dimensions[r].height = src_ws.row_dimensions[r].height


def _safe_value(ws_vals, r, c):
    try:
        return ws_vals.cell(r, c).value
    except Exception:
        return None


def _copy_block_values_only(src_ws, src_ws_vals, dst_ws, src_row_start, src_row_end, dst_row_start, max_col):
    """
    Copy block while:
    - keeping formatting (styles, row heights, merges)
    - writing VALUES (not formulas) using src_ws_vals (data_only=True)
    """
    row_delta = dst_row_start - src_row_start

    for r in range(src_row_start, src_row_end + 1):
        if r in src_ws.row_dimensions and src_ws.row_dimensions[r].height is not None:
            dst_ws.row_dimensions[r + row_delta].height = src_ws.row_dimensions[r].height

        for c in range(1, max_col + 1):
            src_cell = src_ws.cell(r, c)
            dst_cell = dst_ws.cell(r + row_delta, c)

            # style
            if src_cell.has_style:
                dst_cell._style = copy(src_cell._style)
            dst_cell.number_format = src_cell.number_format
            dst_cell.protection = copy(src_cell.protection)
            dst_cell.alignment = copy(src_cell.alignment)
            dst_cell.font = copy(src_cell.font)
            dst_cell.border = copy(src_cell.border)
            dst_cell.fill = copy(src_cell.fill)

            # value (not formula)
            val = _safe_value(src_ws_vals, r, c)
            if val is None:
                val = src_cell.value
                if isinstance(val, str) and val.startswith("="):
                    val = None
            dst_cell.value = val

    # merges fully inside copied block
    for mr in list(src_ws.merged_cells.ranges):
        min_r, min_c, max_r, max_c = mr.min_row, mr.min_col, mr.max_row, mr.max_col
        if min_r >= src_row_start and max_r <= src_row_end and max_c <= max_col:
            dst_ws.merge_cells(
                start_row=min_r + row_delta,
                start_column=min_c,
                end_row=max_r + row_delta,
                end_column=max_c,
            )


def _merge_anchor(ws, r, c):
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= r <= mr.max_row and mr.min_col <= c <= mr.max_col:
            return mr.min_row, mr.min_col
    return r, c


def _freeze_formulas_to_values(ws, ws_vals):
    """
    Kept (for safety) – but master copy is already values-only.
    This just reduces risk of weird formulas propagating inside each source ws before copy.
    """
    max_row = ws.max_row
    done = set()

    for r in range(DATA_START_ROW, max_row + 1):
        for c in FREEZE_COLS:
            ar, ac = _merge_anchor(ws, r, c)
            if (ar, ac) in done:
                continue

            target = ws.cell(ar, ac)
            if isinstance(target, MergedCell):
                continue

            cached = ws_vals.cell(ar, ac).value
            if cached is not None:
                target.value = cached

            done.add((ar, ac))


def _insert_wfp_column_a(ws, ws_vals, data_start, data_end):
    ws.insert_cols(1)
    ws_vals.insert_cols(1)

    style_source = ws.cell(data_start, 2)  # B
    for r in range(data_start, data_end + 1):
        has_data = False
        for c in range(2, ws.max_column + 1):
            if ws.cell(r, c).value not in (None, ""):
                has_data = True
                break
        if not has_data:
            continue

        a_cell = ws.cell(r, 1)
        a_cell.value = "WFP"

        if style_source is not None and style_source.has_style:
            a_cell._style = copy(style_source._style)
            a_cell.number_format = style_source.number_format
            a_cell.protection = copy(style_source.protection)
            a_cell.alignment = copy(style_source.alignment)
            a_cell.font = copy(style_source.font)
            a_cell.border = copy(style_source.border)
            a_cell.fill = copy(style_source.fill)


def _row_has_data_for_merge(ws, r, max_col):
    for c in range(2, max_col + 1):
        if ws.cell(r, c).value not in (None, ""):
            return True
    return False


def _merge_wfp_blocks_in_master(master_ws, max_col):
    max_row = master_ws.max_row

    # unmerge anything intersecting col A in data area
    to_unmerge = []
    for mr in list(master_ws.merged_cells.ranges):
        min_c, min_r, max_c, max_r = range_boundaries(str(mr))
        if min_c <= 1 <= max_c and max_r >= DATA_START_ROW:
            to_unmerge.append(str(mr))
    for rng in to_unmerge:
        master_ws.unmerge_cells(rng)

    r = DATA_START_ROW
    while r <= max_row:
        if master_ws.cell(r, 1).value == "WFP" and _row_has_data_for_merge(master_ws, r, max_col):
            start = r
            end = r
            r += 1
            while (
                r <= max_row
                and master_ws.cell(r, 1).value == "WFP"
                and _row_has_data_for_merge(master_ws, r, max_col)
            ):
                end = r
                r += 1

            if end > start:
                top = master_ws.cell(start, 1)
                top_style = copy(top._style)
                top_font = copy(top.font)
                top_alignment = copy(top.alignment)
                top_border = copy(top.border)
                top_fill = copy(top.fill)
                top_number_format = top.number_format
                top_protection = copy(top.protection)

                master_ws.merge_cells(start_row=start, start_column=1, end_row=end, end_column=1)

                top = master_ws.cell(start, 1)
                top.value = "WFP"
                top._style = top_style
                top.font = top_font
                top.alignment = top_alignment
                top.border = top_border
                top.fill = top_fill
                top.number_format = top_number_format
                top.protection = top_protection
        else:
            r += 1


def _clear_merges_on_cols(ws, cols=(1, 2), from_row=DATA_START_ROW):
    to_unmerge = []

    for mr in list(ws.merged_cells.ranges):
        min_c, min_r, max_c, max_r = range_boundaries(str(mr))

        if max_r < from_row:
            continue

        intersects_cols = any(min_c <= col <= max_c for col in cols)
        if intersects_cols:
            to_unmerge.append(str(mr))

    for rng in to_unmerge:
        ws.unmerge_cells(rng)


def _merge_down_by_blanks(ws, col, max_col, from_row=DATA_START_ROW):
    """
    Merge downwards starting at a value cell, until next value appears,
    as long as the row is a real data row.
    """
    r = from_row
    while r <= ws.max_row:
        v = ws.cell(r, col).value
        if v in (None, "") or not _row_has_data_row(ws, r, max_col):
            r += 1
            continue

        start = r
        end = r
        rr = r + 1

        while rr <= ws.max_row:
            if not _row_has_data_row(ws, rr, max_col):
                break

            next_val = ws.cell(rr, col).value
            if next_val not in (None, ""):
                break

            end = rr
            rr += 1

        if end > start:
            top = ws.cell(start, col)
            top_style = copy(top._style)
            top_font = copy(top.font)
            top_alignment = copy(top.alignment)
            top_border = copy(top.border)
            top_fill = copy(top.fill)
            top_number_format = top.number_format
            top_protection = copy(top.protection)

            ws.merge_cells(start_row=start, start_column=col, end_row=end, end_column=col)

            top = ws.cell(start, col)
            top.value = v
            top._style = top_style
            top.font = top_font
            top.alignment = top_alignment
            top.border = top_border
            top.fill = top_fill
            top.number_format = top_number_format
            top.protection = top_protection

        r = end + 1


def _clear_merges_intersecting_cols(ws, col_start, col_end):
    to_unmerge = []

    for mr in list(ws.merged_cells.ranges):
        min_c, min_r, max_c, max_r = range_boundaries(str(mr))

        if not (max_c < col_start or min_c > col_end):
            to_unmerge.append(str(mr))

    for rng in to_unmerge:
        ws.unmerge_cells(rng)


def _delete_cols_safe(ws, col_start, col_end):
    _clear_merges_intersecting_cols(ws, col_start, col_end)
    for col in range(col_end, col_start - 1, -1):
        if col <= ws.max_column:
            ws.delete_cols(col, 1)

def _ensure_real_cell(ws, row, col):
    """
    Converts a stale MergedCell into a normal writable Cell.
    This is only needed because old code directly edited ws.merged_cells.ranges.
    """
    cell = ws.cell(row=row, column=col)

    if isinstance(cell, MergedCell):
        ws._cells.pop((row, col), None)
        cell = ws.cell(row=row, column=col)

    return cell

def build_combined_workbook_bytes(uploads, status=None):
    """
    Returns (combined_bytes_io, combined_workbook_object).
    """
    tasks = []

    # Load each uploaded workbook twice: normal + data_only cache
    for up in uploads:
        raw = BytesIO(up.getvalue())
        try:
            wb = load_workbook(raw, data_only=False)
            raw.seek(0)
            wb_vals = load_workbook(raw, data_only=True)
        except Exception as e:
            if status:
                status.warning(f"Skipped '{up.name}': could not read as .xlsx ({e})")
            continue

        kind, sheetname = _find_target_sheet(wb)
        if kind is None:
            if status:
                status.warning(f"Skipped '{up.name}': no matching Total Distribution sheet found. Sheets: {wb.sheetnames}")
            continue

        tasks.append({"name": up.name, "kind": kind, "sheet": sheetname, "wb": wb, "wb_vals": wb_vals})

    if not tasks:
        raise RuntimeError("No valid files found.")

    # UNOPS first, then WFP
    tasks.sort(key=lambda x: 0 if x["kind"] == "UNOPS" else 1)

    # Pre-pass: compute bounds + max cols
    max_cols = []
    bounds = []
    for t in tasks:
        ws = t["wb"][t["sheet"]]
        start_row = _first_data_row(ws, DATA_START_ROW)
        end_row = _last_data_row(ws, start_row, ws.max_column)
        bounds.append((start_row, end_row))
        max_cols.append(ws.max_column)

    master_max_col = max(max_cols) if max_cols else 1

    # Master workbook
    master_wb = Workbook()
    master_ws = master_wb.active
    master_ws.title = "Distribution Summary"

    wrote_header = False
    next_write_row = DATA_START_ROW

    for idx, t in enumerate(tasks):
        ws = t["wb"][t["sheet"]]
        ws_vals = t["wb_vals"][t["sheet"]]

        start_row, end_row = bounds[idx]
        if end_row < start_row:
            if status:
                status.warning(f"⚠️ {t['kind']} | {t['name']} has no data rows.")
            continue

        # WFP insert A + stamp WFP on data rows
        if t["kind"] == "WFP":
            _insert_wfp_column_a(ws, ws_vals, start_row, end_row)
            master_max_col = max(master_max_col, ws.max_column)

        # Recompute end_row after any structural change
        end_row = _last_data_row(ws, start_row, ws.max_column)
        if end_row < start_row:
            if status:
                status.warning(f"⚠️ {t['kind']} | {t['name']} has no data rows after WFP insert.")
            continue

        # Delete last text row BEFORE combining
        end_row = _delete_last_text_row_in_sheet(ws, start_row, end_row, ws.max_column)
        if end_row < start_row:
            if status:
                status.warning(f"⚠️ {t['kind']} | {t['name']} became empty after removing last text row.")
            continue

        # Freeze some formulas in-source (optional safety)
        _freeze_formulas_to_values(ws, ws_vals)

        # Copy header once
        if not wrote_header:
            _copy_dimensions(ws, master_ws, max_col=master_max_col, header_rows=HEADER_ROWS)
            _copy_block_values_only(ws, ws_vals, master_ws, 1, HEADER_ROWS, 1, max_col=master_max_col)
            wrote_header = True

        # Copy data values-only
        _copy_block_values_only(ws, ws_vals, master_ws, start_row, end_row, next_write_row, max_col=master_max_col)
        next_write_row += (end_row - start_row + 1)

    # Merge WFP blocks A
    _merge_wfp_blocks_in_master(master_ws, max_col=master_ws.max_column)

    # Rebuild merges A/B by blanks
    _clear_merges_on_cols(master_ws, cols=(1, 2), from_row=DATA_START_ROW)
    _merge_down_by_blanks(master_ws, col=1, max_col=master_ws.max_column, from_row=DATA_START_ROW)
    _merge_down_by_blanks(master_ws, col=2, max_col=master_ws.max_column, from_row=DATA_START_ROW)

    # Ensure visible
    for c in range(1, master_ws.max_column + 1):
        master_ws.column_dimensions[get_column_letter(c)].hidden = False
    for r in range(1, master_ws.max_row + 1):
        master_ws.row_dimensions[r].hidden = False
    master_ws.auto_filter.ref = None

    # Delete F–I on combined (if requested)
    if DELETE_COMBINED_COLS_F_TO_I_BEFORE_CALC:
        _delete_cols_safe(master_ws, 6, 9)
    
    out = BytesIO()
    master_wb.save(out)
    out.seek(0)
    return out


# ============================================================
# CALCULATIONS HELPERS (your cleaner logic, wrapped into a function)
# ============================================================
def copy_cell_style(src, dst):
    dst.font = copy(src.font)
    dst.fill = copy(src.fill)
    dst.border = copy(src.border)
    dst.alignment = copy(src.alignment)
    dst.number_format = src.number_format
    dst.protection = copy(src.protection)


def unmerge_and_fill(ws, col_min: int, col_max: int):
    merges = []
    for mr in ws.merged_cells.ranges:
        if not (mr.max_col < col_min or mr.min_col > col_max):
            merges.append(mr)

    for mr in merges:
        tl = ws.cell(row=mr.min_row, column=mr.min_col)
        value = tl.value
        style = {
            "font": copy(tl.font),
            "fill": copy(tl.fill),
            "border": copy(tl.border),
            "alignment": copy(tl.alignment),
            "number_format": tl.number_format,
            "protection": copy(tl.protection),
        }

        ws.unmerge_cells(str(mr))

        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                cell = ws.cell(row=r, column=c)
                cell.value = value
                cell.font = copy(style["font"])
                cell.fill = copy(style["fill"])
                cell.border = copy(style["border"])
                cell.alignment = copy(style["alignment"])
                cell.number_format = style["number_format"]
                cell.protection = copy(style["protection"])


def snapshot_row(ws, r):
    data = []
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(r, c)
        data.append(
            {
                "value": cell.value,
                "font": copy(cell.font),
                "fill": copy(cell.fill),
                "border": copy(cell.border),
                "alignment": copy(cell.alignment),
                "number_format": cell.number_format,
                "protection": copy(cell.protection),
            }
        )
    return data


def restore_row(ws, r, row_data):
    for idx, d in enumerate(row_data, start=1):
        cell = ws.cell(r, idx)
        cell.value = d["value"]
        cell.font = d["font"]
        cell.fill = d["fill"]
        cell.border = d["border"]
        cell.alignment = d["alignment"]
        cell.number_format = d["number_format"]
        cell.protection = d["protection"]


def safe_float(v):
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip()
        if s == "":
            return 0.0
        s = s.replace(",", "")
        try:
            return float(s)
        except Exception:
            return 0.0
    return 0.0


def norm_header(x):
    return "" if x is None else " ".join(str(x).split()).strip().upper()


def rgb_fill(r, g, b):
    return PatternFill(fill_type="solid", fgColor=f"FF{r:02X}{g:02X}{b:02X}")

def style_headers_black_only_with_text(ws, header_row=1):
    header_fill = PatternFill(fill_type="solid", fgColor="000000")  # black
    header_font = Font(color="FFFFFF", bold=True)                  # white

    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=c)

        # Only style if header has real text
        if cell.value is not None and str(cell.value).strip() != "":
            cell.fill = header_fill
            cell.font = header_font

def _font_without_bold(f: Font) -> Font:
    if f is None:
        return Font(bold=False)
    return Font(
        name=f.name,
        sz=f.sz,
        b=False,
        i=f.italic,
        u=f.underline,
        strike=f.strike,
        color=f.color,
        vertAlign=f.vertAlign,
        outline=f.outline,
        shadow=f.shadow,
        condense=f.condense,
        extend=f.extend,
        charset=f.charset,
        family=f.family,
        scheme=f.scheme,
    )


def remove_bold_except_header(ws, header_row=1):
    """
    Remove bold from all cells except header row.
    Keeps other font properties intact.
    """
    max_r = ws.max_row
    max_c = ws.max_column
    for r in range(1, max_r + 1):
        if r == header_row:
            continue
        for c in range(1, max_c + 1):
            cell = ws.cell(row=r, column=c)
            f = cell.font
            if f is not None and f.bold:
                cell.font = _font_without_bold(f)


def _find_header_col_by_names(ws, header_row, header_names):
    """Find a column by one of several header names using the existing normaliser."""
    wanted = {norm_header(name) for name in header_names}
    for c in range(1, ws.max_column + 1):
        if norm_header(ws.cell(row=header_row, column=c).value) in wanted:
            return c
    return None


def _strip_leading_usage_marker(value):
    """
    New UNOPS format can prefix Description values with:
      - Internal use -
      - external use
    Return (clean_value, usage_type). usage_type is 'internal', 'external', or None.
    """
    if value is None:
        return value, None

    text_value = str(value)
    pattern = re.compile(r"^\s*(internal\s+use|external\s+use)\b\s*[-–—:]?\s*", re.IGNORECASE)
    match = pattern.match(text_value)
    if not match:
        return value, None

    usage_type = "internal" if match.group(1).strip().lower().startswith("internal") else "external"
    clean_value = pattern.sub("", text_value, count=1).strip()
    return clean_value, usage_type


def _compact_alpha_num(value) -> str:
    if value is None:
        return ""
    return re.sub(r"[^A-Z0-9]", "", str(value).upper())


def _contains_ohchr(value) -> bool:
    return "OHCHR" in _compact_alpha_num(value)


def _is_regular_cluster_for_internal_use(value) -> bool:
    """
    Internal-use rows should become LOGISTICS only when the original cluster is a
    regular cluster. If the cluster is irregular (for example, UN - Sisters Agencies),
    keep it untouched so the existing irregular-cluster logic converts it to INGOs
    with the generated prefix: <cluster> - <agency>.
    """
    return norm_header(value) in {
        "ETC",
        "HEALTH",
        "WASH",
        "INGOS",
        "UN-OHCHR",
        "UN OHCHR",
        "UN AGENCIES",
        "LOGISTICS",
    }


def normalise_distribution_description_usage_markers(ws, header_row=1):
    """
    Normalise the new UNOPS Description format back to the old working format.

    - Removes leading 'Internal use -' / 'external use' from Description.
    - Internal-use rows become LOGISTICS only when the original cluster is regular.
      Irregular clusters stay untouched so the existing logic later converts them
      to INGOs and prefixes the Agency with the original cluster.
    - OHCHR rows are always treated as LOGISTICS so they later become UN Agencies,
      not a generated irregular entry such as 'Protection - OHCHR'.
    - WFP rows are left as WFP to preserve WFP-specific handling.
    """
    description_col = _find_header_col_by_names(ws, header_row, ["DESCRIPTION", "Description"])
    cluster_col = _find_header_col_by_names(ws, header_row, ["CLUSTER", "Cluster", "INTERVENTION", "Intervention"])
    agency_col = _find_header_col_by_names(ws, header_row, ["AGENCY", "Agency"])

    if description_col is None:
        return

    for r in range(header_row + 1, ws.max_row + 1):
        cluster_cell = ws.cell(row=r, column=cluster_col) if cluster_col is not None else None
        agency_value = ws.cell(row=r, column=agency_col).value if agency_col is not None else None
        cluster_value = cluster_cell.value if cluster_cell is not None else None

        if cluster_cell is not None and (_contains_ohchr(cluster_value) or _contains_ohchr(agency_value)):
            cluster_cell.value = "LOGISTICS"
            cluster_value = cluster_cell.value

        desc_cell = ws.cell(row=r, column=description_col)
        cleaned_desc, usage_type = _strip_leading_usage_marker(desc_cell.value)

        if usage_type is None:
            continue

        desc_cell.value = cleaned_desc

        if usage_type == "internal" and cluster_cell is not None:
            cluster_norm = norm_header(cluster_value)

            # Do not convert WFP rows; the new marker is intended for UNOPS Total Distribution.
            # Irregular clusters must keep the original cluster so the old irregular-cluster
            # conversion still produces: <cluster> - <agency>, then Cluster = INGOs.
            if cluster_norm != "WFP" and _is_regular_cluster_for_internal_use(cluster_value):
                cluster_cell.value = "LOGISTICS"

def run_calculations_on_combined_bytes(combined_bytes: BytesIO, progress=None, status=None) -> BytesIO:
    """
    Takes the combined workbook bytes (with sheet 'Master'),
    runs your calculations pipeline on it,
    returns final output BytesIO.
    """
    combined_bytes.seek(0)
    wb = load_workbook(combined_bytes, data_only=False)
    combined_bytes.seek(0)
    wb_cache = load_workbook(combined_bytes, data_only=True)

    # We already have only one sheet: "Master"
    TARGET_SHEET = "Distribution Summary"
    sheet_map = {name.strip(): name for name in wb.sheetnames}
    sheet_map_cache = {name.strip(): name for name in wb_cache.sheetnames}

    if TARGET_SHEET not in sheet_map or TARGET_SHEET not in sheet_map_cache:
        raise RuntimeError(f'Sheet "{TARGET_SHEET}" not found. Sheets found: {list(sheet_map.keys())}')

    keep_name = sheet_map[TARGET_SHEET]
    keep_name_cache = sheet_map_cache[TARGET_SHEET]

    # keep only Master in both (defensive)
    for name in list(wb.sheetnames):
        if name != keep_name:
            wb.remove(wb[name])
    for name in list(wb_cache.sheetnames):
        if name != keep_name_cache:
            wb_cache.remove(wb_cache[name])

    ws = wb[keep_name]
    ws_cache = wb_cache[keep_name_cache]
    ws.title = TARGET_SHEET

    if progress:
        progress.progress(10)
    if status:
        status.info("Unmerging A–C…")

    # Unmerge A–C
    unmerge_and_fill(ws, col_min=1, col_max=3)

    if progress:
        progress.progress(18)
    if status:
        status.info("Freezing formulas into values…")

    # Freeze any remaining formulas into values using cache
    max_row = ws.max_row
    max_col = ws.max_column
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            is_formula = (cell.data_type == "f") or (isinstance(cell.value, str) and cell.value.startswith("="))
            if is_formula:
                cell.value = ws_cache.cell(row=r, column=c).value

    if progress:
        progress.progress(26)

    # Delete row 2
    ws.delete_rows(2)

    # Delete rows containing 'TOTAL' in A,B,C (kept from your original)
    if status:
        status.info("Removing TOTAL rows…")

    rows_to_delete = []
    for row in range(1, ws.max_row + 1):
        for col in (1, 2, 3):
            val = ws.cell(row=row, column=col).value
            if val is not None and "TOTAL" in str(val).upper():
                rows_to_delete.append(row)
                break
    for rr in sorted(set(rows_to_delete), reverse=True):
        ws.delete_rows(rr)

    if progress:
        progress.progress(34)

    header_row = 1

    # Normalise new UNOPS Description prefixes before the existing calculation logic runs.
    normalise_distribution_description_usage_markers(ws, header_row=header_row)

    # Unmerge D–F before Fuel Sum
    unmerge_and_fill(ws, col_min=4, col_max=6)  # D..F

    # Add Fuel sum in F (values)
    if status:
        status.info("Building Fuel sum…")

    col_d, col_e, col_f = 4, 5, 6  # D,E,F

    ref_header = ws.cell(row=header_row, column=col_d)
    fuel_header = ws.cell(row=header_row, column=col_f)

    fuel_header.value = "Fuel sum"
    fuel_header.font = copy(ref_header.font)
    fuel_header.fill = copy(ref_header.fill)
    fuel_header.border = copy(ref_header.border)
    fuel_header.alignment = copy(ref_header.alignment)
    fuel_header.number_format = ref_header.number_format
    fuel_header.protection = copy(ref_header.protection)

    if "E" in ws.column_dimensions and ws.column_dimensions["E"].width is not None:
        ws.column_dimensions["F"].width = ws.column_dimensions["E"].width
    elif "D" in ws.column_dimensions and ws.column_dimensions["D"].width is not None:
        ws.column_dimensions["F"].width = ws.column_dimensions["D"].width

    for r in range(header_row + 1, ws.max_row + 1):
        d_cell = ws.cell(row=r, column=col_d)
        e_cell = ws.cell(row=r, column=col_e)
        f_cell = ws.cell(row=r, column=col_f)

        d_val = d_cell.value
        e_val = e_cell.value

        if d_val is None and e_val is None:
            f_cell.value = None
        else:
            try:
                d_num = 0.0 if d_val is None else float(d_val)
                e_num = 0.0 if e_val is None else float(e_val)
                f_cell.value = d_num + e_num
            except Exception:
                f_cell.value = None

        style_src = d_cell if d_cell.value is not None else e_cell
        f_cell.font = copy(style_src.font)
        f_cell.fill = copy(style_src.fill)
        f_cell.border = copy(style_src.border)
        f_cell.alignment = copy(style_src.alignment)
        f_cell.number_format = style_src.number_format
        f_cell.protection = copy(style_src.protection)

    if progress:
        progress.progress(45)

    # Delete rows where Fuel sum is 0 or empty
    if status:
        status.info("Removing empty/zero fuel rows…")

    rows_remove = []
    for r in range(header_row + 1, ws.max_row + 1):
        val = ws.cell(row=r, column=col_f).value
        if val is None or (isinstance(val, str) and val.strip() == ""):
            rows_remove.append(r)
            continue
        try:
            if float(val) == 0.0:
                rows_remove.append(r)
        except Exception:
            pass
    for rr in sorted(set(rows_remove), reverse=True):
        ws.delete_rows(rr)

    if progress:
        progress.progress(52)

    # Delete columns D and E
    ws.delete_cols(5)  # E
    ws.delete_cols(4)  # D
    # Fuel sum moved to D.

    if progress:
        progress.progress(56)

    # Insert Description Sum as column E
    if status:
        status.info("Building Description Sum…")

    desc_col = 5  # E
    ws.insert_cols(desc_col)

    ref_h = ws.cell(row=header_row, column=1)
    desc_h = ws.cell(row=header_row, column=desc_col)

    desc_h.value = "Description Sum"
    desc_h.font = copy(ref_h.font)
    desc_h.fill = copy(ref_h.fill)
    desc_h.border = copy(ref_h.border)
    desc_h.alignment = copy(ref_h.alignment)
    desc_h.number_format = ref_h.number_format
    desc_h.protection = copy(ref_h.protection)

    if "C" in ws.column_dimensions and ws.column_dimensions["C"].width is not None:
        ws.column_dimensions[get_column_letter(desc_col)].width = ws.column_dimensions["C"].width

    for r in range(header_row + 1, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        b = ws.cell(row=r, column=2).value
        c = ws.cell(row=r, column=3).value

        a_s = "" if a is None else str(a)
        b_s = "" if b is None else str(b)
        c_s = "" if c is None else str(c)

        cell = ws.cell(row=r, column=desc_col)
        cell.value = f"{a_s},{b_s},{c_s}"

        src = ws.cell(row=r, column=1)
        cell.font = copy(src.font)
        cell.fill = copy(src.fill)
        cell.border = copy(src.border)
        cell.alignment = copy(src.alignment)
        cell.number_format = src.number_format
        cell.protection = copy(src.protection)

    if progress:
        progress.progress(64)

    # Unified Fuel as values
    if status:
        status.info("Building Unified Fuel…")

    unified_col = 6  # F
    ws.insert_cols(unified_col)

    ref_unified_header = ws.cell(row=header_row, column=4)  # D header
    unified_header = ws.cell(row=header_row, column=unified_col)
    unified_header.value = "Unified Fuel"
    unified_header.font = copy(ref_unified_header.font)
    unified_header.fill = copy(ref_unified_header.fill)
    unified_header.border = copy(ref_unified_header.border)
    unified_header.alignment = copy(ref_unified_header.alignment)
    unified_header.number_format = ref_unified_header.number_format
    unified_header.protection = copy(ref_unified_header.protection)

    d_letter = get_column_letter(4)
    f_letter = get_column_letter(unified_col)
    if d_letter in ws.column_dimensions and ws.column_dimensions[d_letter].width is not None:
        ws.column_dimensions[f_letter].width = ws.column_dimensions[d_letter].width

    totals = {}
    for r in range(header_row + 1, ws.max_row + 1):
        key = ws.cell(row=r, column=5).value  # E
        fuel = safe_float(ws.cell(row=r, column=4).value)  # D
        totals[key] = totals.get(key, 0.0) + fuel

    for r in range(header_row + 1, ws.max_row + 1):
        target = ws.cell(row=r, column=unified_col)  # F
        style_src = ws.cell(row=r, column=4)         # D
        key = ws.cell(row=r, column=5).value         # E

        target.value = totals.get(key, 0.0)
        target.font = copy(style_src.font)
        target.fill = copy(style_src.fill)
        target.border = copy(style_src.border)
        target.alignment = copy(style_src.alignment)
        target.number_format = style_src.number_format
        target.protection = copy(style_src.protection)

    if progress:
        progress.progress(72)

    # Sorting
    if status:
        status.info("Sorting…")

    # Find INTERVENTION + AGENCY columns
    intervention_col = None
    agency_col = None

    for c in range(1, ws.max_column + 1):
        h = ws.cell(row=1, column=c).value
        nh = norm_header(h)
        if nh == "INTERVENTION":
            intervention_col = c
        elif nh == "AGENCY":
            agency_col = c

    if intervention_col is None:
        intervention_col = 1

    if agency_col is None:
        raise RuntimeError('Header "AGENCY" not found.')

    # Normalise OHCHR rows to LOGISTICS so they later become UN Agencies,
    # even when the source file places OHCHR under an irregular cluster such as Protection.
    for r in range(2, ws.max_row + 1):
        iv = ws.cell(row=r, column=intervention_col).value
        agency = ws.cell(row=r, column=agency_col).value
        if _contains_ohchr(iv) or _contains_ohchr(agency):
            ws.cell(row=r, column=intervention_col).value = "LOGISTICS"

    REGULAR_INTERVENTIONS = {
        "ETC",
        "HEALTH",
        "WASH",
        "INGOS",
        "UN-OHCHR",
        "WFP",
        "LOGISTICS",   # IMPORTANT: exclude from prefix+convert logic
    }

    def _clean_str(v):
        return "" if v is None else str(v).strip()

    converted_desc_keys = set()  # <-- stable IDs for converted rows
    rows_data = []
    max_r = ws.max_row  # cache for speed

    for r in range(2, max_r + 1):
        iv_raw = _clean_str(ws.cell(row=r, column=intervention_col).value)
        iv_up = iv_raw.upper()

        is_converted = False

        # Apply your rule ONLY when intervention is not regular
        if iv_raw and (iv_up not in REGULAR_INTERVENTIONS):
            agency_cell = ws.cell(row=r, column=agency_col)
            agency_raw = _clean_str(agency_cell.value)

            prefix = f"{iv_raw} - "
            if not agency_raw.startswith(prefix):
                agency_cell.value = f"{iv_raw} - {agency_raw}" if agency_raw else f"{iv_raw} -"

            # Convert INTERVENTION to INGOs
            ws.cell(row=r, column=intervention_col).value = "INGOs"
            is_converted = True

        # After any conversion, read the final values used for sorting + tracking
        fuel_val = safe_float(ws.cell(row=r, column=unified_col).value)
        iv_after = _clean_str(ws.cell(row=r, column=intervention_col).value)
        desc_key = ws.cell(row=r, column=desc_col).value

        # Track converted rows by stable key (survives sorting/deletions)
        if is_converted:
            converted_desc_keys.add(desc_key)

        rows_data.append({
            "fuel": fuel_val,
            "intervention": iv_after,
            "is_converted": is_converted,
            "orig_index": r,
            "desc_key": desc_key,
            "row": snapshot_row(ws, r),
        })


    # ONE sort (tuple key) so behaviour is deterministic:
    # 1) intervention A–Z
    # 2) inside INGOs: real INGOs first, converted/prefixed last
    # 3) fuel DESC only for rows that are NOT converted prefixed INGOs
    # 4) stable fallback by orig_index
    def _sort_key(x):
        iv = (x["intervention"] or "").strip().lower()
        is_ingos = (iv == "ingos")

        # 0 = purple (real INGOs), 1 = white (converted/prefixed)
        converted_rank = 1 if (is_ingos and x["is_converted"]) else 0

        # Sort fuel DESC inside BOTH purple and white groups (independently)
        fuel_rank = -x["fuel"] if is_ingos else -x["fuel"]

        return (iv, converted_rank, fuel_rank, x["orig_index"])

    rows_data.sort(key=_sort_key)

    # Restore rows and rebuild "converted rows" positions AFTER sorting
    write_row = 2
    for obj in rows_data:
        restore_row(ws, write_row, obj["row"])
        write_row += 1


    if progress:
        progress.progress(80)

    # Remove duplicates by Description Sum
    if status:
        status.info("Removing duplicates…")

    seen = set()
    dup_rows = []
    for r in range(2, ws.max_row + 1):
        val = ws.cell(row=r, column=desc_col).value
        key = "" if val is None else str(val).strip()
        if key in seen:
            dup_rows.append(r)
        else:
            seen.add(key)
    for rr in sorted(dup_rows, reverse=True):
        ws.delete_rows(rr)

    if progress:
        progress.progress(86)

    # Total Sum Per Category in G
    if status:
        status.info("Building Total Sum Per Category…")

    total_cat_col = 7  # G
    if ws.max_column >= total_cat_col:
        ws.insert_cols(total_cat_col)

    ref_total_header = ws.cell(row=header_row, column=unified_col)  # F header
    total_header = ws.cell(row=header_row, column=total_cat_col)    # G header
    total_header.value = "Total Sum Per Category"
    total_header.font = copy(ref_total_header.font)
    total_header.fill = copy(ref_total_header.fill)
    total_header.border = copy(ref_total_header.border)
    total_header.alignment = copy(ref_total_header.alignment)
    total_header.number_format = ref_total_header.number_format
    total_header.protection = copy(ref_total_header.protection)

    ref_letter = get_column_letter(unified_col)
    new_letter = get_column_letter(total_cat_col)
    if ref_letter in ws.column_dimensions and ws.column_dimensions[ref_letter].width is not None:
        ws.column_dimensions[new_letter].width = ws.column_dimensions[ref_letter].width

    cat_totals = {}
    for r in range(2, ws.max_row + 1):
        cat = ws.cell(row=r, column=1).value
        cat_key = "" if cat is None else str(cat).strip()
        fuel_val = safe_float(ws.cell(row=r, column=unified_col).value)
        cat_totals[cat_key] = cat_totals.get(cat_key, 0.0) + fuel_val

    for r in range(2, ws.max_row + 1):
        cat = ws.cell(row=r, column=1).value
        cat_key = "" if cat is None else str(cat).strip()

        target = ws.cell(row=r, column=total_cat_col)
        style_src = ws.cell(row=r, column=unified_col)

        target.value = cat_totals.get(cat_key, 0.0)
        target.font = copy(style_src.font)
        target.fill = copy(style_src.fill)
        target.border = copy(style_src.border)
        target.alignment = copy(style_src.alignment)
        target.number_format = style_src.number_format
        target.protection = copy(style_src.protection)

    if progress:
        progress.progress(92)

    # Merge Total Sum Per Category (G) by INTERVENTION
    if status:
        status.info('Merging "Total Sum Per Category" by INTERVENTION…')

    INTERVENTION_COL = intervention_col
    TOTAL_CAT_COL = total_cat_col

    def _norm_intervention(v):
        return "" if v is None else str(v).strip().upper()

    start = 2
    while start <= ws.max_row:
        key = _norm_intervention(ws.cell(row=start, column=INTERVENTION_COL).value)

        end = start
        while end + 1 <= ws.max_row and _norm_intervention(ws.cell(row=end + 1, column=INTERVENTION_COL).value) == key:
            end += 1

        if end > start:
            top_cell = ws.cell(row=start, column=TOTAL_CAT_COL)
            top_val = top_cell.value
            top_style = {
                "font": copy(top_cell.font),
                "fill": copy(top_cell.fill),
                "border": copy(top_cell.border),
                "alignment": copy(top_cell.alignment),
                "number_format": top_cell.number_format,
                "protection": copy(top_cell.protection),
            }

            ws.merge_cells(
                start_row=start,
                start_column=TOTAL_CAT_COL,
                end_row=end,
                end_column=TOTAL_CAT_COL,
            )

            merged_top = ws.cell(row=start, column=TOTAL_CAT_COL)
            merged_top.value = top_val
            merged_top.font = copy(top_style["font"])
            merged_top.fill = copy(top_style["fill"])
            merged_top.border = copy(top_style["border"])
            merged_top.alignment = copy(top_style["alignment"])
            merged_top.number_format = top_style["number_format"]
            merged_top.protection = copy(top_style["protection"])

        start = end + 1

    # OPTIONAL: rename LOGISTICS -> UN Agencies
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=intervention_col).value
        if v is None:
            continue
        if str(v).strip().upper() == "LOGISTICS":
            ws.cell(row=r, column=intervention_col).value = "UN Agencies"

    # Color A–G based on INTERVENTION
    if status:
        status.info("Applying colours…")

    def _is_converted_row(r: int) -> bool:
        iv = ws.cell(row=r, column=intervention_col).value
        if ("" if iv is None else str(iv).strip().upper()) != "INGOS":
            return False
        return ws.cell(row=r, column=desc_col).value in converted_desc_keys

    fills = {
        "ETC": rgb_fill(213, 243, 251),
        "HEALTH": rgb_fill(0, 176, 80),
        "WASH": rgb_fill(250, 178, 138),
        "INGOs": rgb_fill(190, 158, 242),
        "WFP": rgb_fill(44, 195, 236),
        "UN_AGENCIES": rgb_fill(0, 176, 240),
    }

    COLOR_MIN_COL = 1
    COLOR_MAX_COL = 7

    for r in range(2, ws.max_row + 1):
        intervention_val = ws.cell(row=r, column=intervention_col).value
        intervention_text = "" if intervention_val is None else str(intervention_val).strip()
        intervention_up = intervention_text.upper()

        row_fill = None

        if intervention_up == "WFP":
            row_fill = fills["WFP"]
        elif intervention_up == "UN AGENCIES":
            row_fill = fills["UN_AGENCIES"]
        elif intervention_up == "ETC":
            row_fill = fills["ETC"]
        elif intervention_up == "HEALTH":
            row_fill = fills["HEALTH"]
        elif intervention_up == "WASH":
            row_fill = fills["WASH"]
        elif intervention_up == "INGOS":
            # Converted/prefixed rows must remain white; real INGOs must be purple
            if _is_converted_row(r):
                row_fill = None
            else:
                row_fill = fills["INGOs"]


        # CRITICAL: never write None into .fill
        if row_fill is None:
            continue

        for c in range(COLOR_MIN_COL, COLOR_MAX_COL + 1):
            ws.cell(row=r, column=c).fill = row_fill


    # Summary sheet (Sector Summary) + pie chart
    totals_by_intervention = {}
    for r in range(2, ws.max_row + 1):
        iv = ws.cell(row=r, column=intervention_col).value
        iv_key = "" if iv is None else str(iv).strip()
        if iv_key == "":
            continue
        totals_by_intervention[iv_key.upper()] = totals_by_intervention.get(iv_key.upper(), 0.0) + safe_float(
            ws.cell(row=r, column=unified_col).value
        )

    sector_rows = [
        ("ETC", "תקשורת"),
        ("HEALTH", "בריאות"),
        ("WASH", "סניטציה"),
        ("INGOS", "ארגונים לא ממשלתיים"),
        ("UN AGENCIES", 'סוכנויות או"ם'),
        ("WFP", "WFP"),
    ]

    SUMMARY_SHEET_NAME = "Sector Summary"
    if SUMMARY_SHEET_NAME in wb.sheetnames:
        wb.remove(wb[SUMMARY_SHEET_NAME])
    ws_sum = wb.create_sheet(SUMMARY_SHEET_NAME)

    ws_sum["A1"].value = "סקטור"
    ws_sum["B1"].value = "כמות דלק (ליטר)"

    header_fill = PatternFill("solid", fgColor="000000")
    header_font = Font(color="FFFFFF", bold=True)
    for cell_ref in ("A1", "B1", "C1"):
        cell = ws_sum[cell_ref]
        cell.fill = header_fill
        cell.font = header_font

    hdr_src = ws.cell(row=1, column=1)
    for addr in ("A1", "B1"):
        c = ws_sum[addr]
        c.font = copy(hdr_src.font)
        c.fill = copy(hdr_src.fill)
        c.border = copy(hdr_src.border)
        c.alignment = copy(hdr_src.alignment)
        c.number_format = hdr_src.number_format
        c.protection = copy(hdr_src.protection)

    ws_sum.column_dimensions["A"].width = 28
    ws_sum.column_dimensions["B"].width = 18

    num_src = ws.cell(row=2, column=unified_col)
    litres_number_format = num_src.number_format

    summary_fills = {
        "ETC": fills["ETC"],
        "HEALTH": fills["HEALTH"],
        "WASH": fills["WASH"],
        "INGOS": fills["INGOs"],
        "UN AGENCIES": fills.get("UN_AGENCIES", rgb_fill(0, 176, 240)),
        "WFP": fills["WFP"],
    }

    style_row_by_intervention = {}
    for r in range(2, ws.max_row + 1):
        iv = ws.cell(row=r, column=intervention_col).value
        k = "" if iv is None else str(iv).strip().upper()
        if k and k not in style_row_by_intervention:
            style_row_by_intervention[k] = r

    row_i = 2
    for key_en, label_he in sector_rows:
        key_u = key_en.strip().upper()

        src_r = style_row_by_intervention.get(key_u)

        a_cell = ws_sum.cell(row=row_i, column=1)
        b_cell = ws_sum.cell(row=row_i, column=2)

        a_cell.value = label_he
        b_cell.value = totals_by_intervention.get(key_u, 0.0)

        if src_r is not None:
            src_a = ws.cell(row=src_r, column=intervention_col)
            src_b = ws.cell(row=src_r, column=unified_col)
            copy_cell_style(src_a, a_cell)
            copy_cell_style(src_b, b_cell)
        else:
            copy_cell_style(hdr_src, a_cell)
            copy_cell_style(hdr_src, b_cell)
            b_cell.number_format = litres_number_format

        fill_obj = summary_fills.get(key_u)
        if fill_obj:
            a_cell.fill = fill_obj
            b_cell.fill = fill_obj

        row_i += 1

    FIRST_ROW = 2
    LAST_ROW = FIRST_ROW + len(sector_rows) - 1

    ws_sum["C1"].value = "אחוז"
    copy_cell_style(ws_sum["A1"], ws_sum["C1"])
    ws_sum.column_dimensions["C"].width = 12

    for r in range(FIRST_ROW, LAST_ROW + 1):
        ws_sum.cell(row=r, column=3).value = f"=B{r}/SUM($B${FIRST_ROW}:$B${LAST_ROW})"
        copy_cell_style(ws_sum.cell(row=r, column=2), ws_sum.cell(row=r, column=3))
        ws_sum.cell(row=r, column=3).number_format = "0.0%"

    pie = PieChart()
    data = Reference(ws_sum, min_col=2, min_row=1, max_row=LAST_ROW)
    cats = Reference(ws_sum, min_col=1, min_row=FIRST_ROW, max_row=LAST_ROW)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(cats)

    pie.title = ""
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showCatName = True
    pie.dataLabels.showPercent = True
    pie.dataLabels.showVal = False
    pie.dataLabels.showSerName = False
    pie.dataLabels.showLeaderLines = True
    pie.dataLabels.separator = "\n"
    pie.legend = None

    slice_colors = [
        "D5F3FB",
        "00B050",
        "FAB28A",
        "BE9EF2",
        "00B0F0",
        "2CC3EC",
    ]

    ser = pie.series[0]
    ser.dPt = []
    for i, hx in enumerate(slice_colors[: len(sector_rows)]):
        dp = DataPoint(idx=i)
        dp.graphicalProperties.solidFill = hx
        ser.dPt.append(dp)

    ws_sum.add_chart(pie, "E2")

    if progress:
        progress.progress(100)

    # Remove ALL bold on Distribution Summary except header row
    remove_bold_except_header(ws, header_row=1)

    # Force header style on both sheets
    style_headers_black_only_with_text(ws, header_row=1)      # Distribution Summary
    style_headers_black_only_with_text(ws_sum, header_row=1)  # Sector Summary

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


# ============================================================
# FUEL DASHBOARD HELPERS (added from standalone Fuel Overview app)
# ============================================================

def find_cell_exact(ws, text: str):
    target = text.strip().upper()
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            value = ws.cell(r, c).value
            if value is None:
                continue
            if str(value).strip().upper() == target:
                return r, c
    return None, None


def find_cell_containing(ws, text: str):
    target = text.upper()
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            value = ws.cell(r, c).value
            if value is not None and target in str(value).upper():
                return r, c
    return None, None


def normalise_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        s = v.strip()
        # In these files, 05/11/2026 means May 11, 2026.
        for fmt in ("%d-%b-%Y", "%m/%d/%Y", "%d/%m/%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(s, fmt).date()
            except Exception:
                pass
    return None


def get_sheet_by_normalised_name(wb, wanted_names):
    wanted = {x.strip().upper() for x in wanted_names}
    for name in wb.sheetnames:
        if name.strip().upper() in wanted:
            return wb[name]
    return None


def unmerge_and_fill_all(ws):
    for mr in list(ws.merged_cells.ranges):
        value = ws.cell(mr.min_row, mr.min_col).value
        ws.unmerge_cells(str(mr))
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                ws.cell(r, c).value = value


def write_number_or_blank(cell, value):
    if value is None or value == 0:
        cell.value = None
    else:
        cell.value = value
        cell.number_format = "#,##0"


def add_log(logs, message):
    if logs is not None:
        logs.append(message)


def cell_ref(row, col):
    return f"{get_column_letter(col)}{row}"


def detect_org(filename: str) -> str:
    name = filename.upper()
    if "UNOPS" in name:
        return "UNOPS"
    if "WFP" in name:
        return "WFP"
    return "Unknown"


def extract_current_fuel_from_file(uploaded_file):
    org = detect_org(uploaded_file.name)
    raw = BytesIO(uploaded_file.getvalue())
    wb = load_workbook(raw, data_only=True)

    if "Summary" not in wb.sheetnames:
        raise RuntimeError(f'"{uploaded_file.name}" does not contain a Summary sheet.')

    ws = wb["Summary"]

    # Fill merged cells only horizontally to avoid duplicating WFP station names vertically.
    for mr in list(ws.merged_cells.ranges):
        value = ws.cell(mr.min_row, mr.min_col).value
        ws.unmerge_cells(str(mr))
        for c in range(mr.min_col, mr.max_col + 1):
            ws.cell(mr.min_row, c).value = value

    header_row, station_col = find_cell_containing(ws, "Fuel Station - Address")
    if not header_row:
        raise RuntimeError(f'"Fuel Station - Address" not found in "{uploaded_file.name}".')

    diesel_col = None
    petrol_col = None
    for c in range(station_col + 1, ws.max_column + 1):
        val = ws.cell(header_row, c).value
        txt = "" if val is None else str(val).strip().upper()
        if txt == "DIESEL":
            diesel_col = c
        elif txt == "PETROL":
            petrol_col = c
        if diesel_col and petrol_col:
            break

    if diesel_col is None:
        raise RuntimeError(f'"Diesel" column not found in "{uploaded_file.name}".')
    if petrol_col is None:
        raise RuntimeError(f'"Petrol" column not found in "{uploaded_file.name}".')

    rows = []
    seen = set()
    for r in range(header_row + 1, ws.max_row + 1):
        row_text = " ".join(
            str(ws.cell(r, c).value)
            for c in range(1, ws.max_column + 1)
            if ws.cell(r, c).value is not None
        ).upper()

        if "CURRENT FUEL IN STORAGE" in row_text:
            break

        station = ws.cell(r, station_col).value
        if station is None or str(station).strip() == "":
            continue

        station_name = str(station).strip()
        diesel = safe_float(ws.cell(r, diesel_col).value)
        petrol = safe_float(ws.cell(r, petrol_col).value)
        key = (org, station_name, diesel, petrol)
        if key in seen:
            continue
        seen.add(key)

        rows.append({
            "Organisation": org,
            "Fuel Station - Address": station_name,
            "Diesel": diesel,
            "Petrol": petrol,
            "Total Fuel": diesel + petrol,
        })

    return rows


def extract_daily_fuel_entries(uploaded_file, selected_dates):
    org = detect_org(uploaded_file.name)
    raw = BytesIO(uploaded_file.getvalue())
    wb = load_workbook(raw, data_only=True)

    ws = get_sheet_by_normalised_name(wb, ["Comulative Fuel Entry", "Cumulative Fuel Entry"])
    if ws is None:
        raise RuntimeError(f'"Comulative Fuel Entry" not found in "{uploaded_file.name}".')

    # Fill merged cells horizontally only, so merged headers like "Note" are readable
    # without wrongly duplicating values down the sheet.
    for mr in list(ws.merged_cells.ranges):
        value = ws.cell(mr.min_row, mr.min_col).value
        ws.unmerge_cells(str(mr))
        for c in range(mr.min_col, mr.max_col + 1):
            ws.cell(mr.min_row, c).value = value

    date_row, date_col = find_cell_containing(ws, "Date")
    diesel_row, diesel_col = find_cell_containing(ws, "Total Diesel Received")
    benzene_row, benzene_col = find_cell_containing(ws, "Total Benzene Received")

    if not date_row:
        raise RuntimeError(f'"Date" not found in "{uploaded_file.name}".')
    if not diesel_row:
        raise RuntimeError(f'"Total Diesel Received" not found in "{uploaded_file.name}".')
    if not benzene_row:
        raise RuntimeError(f'"Total Benzene Received" not found in "{uploaded_file.name}".')

    # Find the Note column near the Diesel/Benzene amount columns.
    # In the WFP file, this Note header is merged with row 2, so we scan the nearby header area.
    note_col = None
    if org == "WFP":
        min_header_row = max(1, min(date_row, diesel_row, benzene_row) - 2)
        max_header_row = min(ws.max_row, max(date_row, diesel_row, benzene_row) + 2)

        # Prefer columns close to Diesel/Benzene
        search_start_col = max(1, min(diesel_col, benzene_col) - 2)
        search_end_col = min(ws.max_column, max(diesel_col, benzene_col) + 5)

        for r in range(min_header_row, max_header_row + 1):
            for c in range(search_start_col, search_end_col + 1):
                val = ws.cell(r, c).value
                txt = "" if val is None else str(val).strip().upper()
                if txt == "NOTE":
                    note_col = c
                    break
            if note_col:
                break

    selected_dates = set(selected_dates)
    results = {}

    for r in range(date_row + 1, ws.max_row + 1):
        current_date = normalise_date(ws.cell(r, date_col).value)
        if current_date is None or current_date not in selected_dates:
            continue

        # WFP-only exclusion:
        # If Note says "From UNOPS", ignore this row completely.
        if org == "WFP" and note_col is not None:
            note_value = ws.cell(r, note_col).value
            note_text = "" if note_value is None else str(note_value).strip().upper()

            if "FROM UNOPS" in note_text:
                continue

        diesel = safe_float(ws.cell(r, diesel_col).value)
        benzene = safe_float(ws.cell(r, benzene_col).value)

        # Add to existing date total instead of replacing it,
        # in case the same date appears more than once.
        results[current_date] = results.get(current_date, 0.0) + diesel + benzene

    return org, results

def extract_fuel_used(uploaded_file, selected_dates, logs=None):
    org = detect_org(uploaded_file.name)
    add_log(logs, f"========== Fuel Used | {uploaded_file.name} | detected org: {org} ==========")

    raw = BytesIO(uploaded_file.getvalue())
    wb = load_workbook(raw, data_only=True)
    ws = get_sheet_by_normalised_name(
        wb,
        [
            "Comulative Ditribution Summary",
            "Comulative Distribution Summary",
            "Cumulative Ditribution Summary",
            "Cumulative Distribution Summary",
        ],
    )
    if ws is None:
        raise RuntimeError(f'"Comulative Ditribution Summary" not found in "{uploaded_file.name}".')

    date_row, date_col = find_cell_exact(ws, "Date")
    sum_row, sum_col = find_cell_exact(ws, "Sum")
    if not date_row:
        raise RuntimeError(f'Exact cell "Date" not found in "{uploaded_file.name}".')
    if not sum_row:
        raise RuntimeError(f'Exact cell "Sum" not found in "{uploaded_file.name}".')

    sum_start_col = sum_col
    sum_end_col = sum_col
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row <= sum_row <= mr.max_row and mr.min_col <= sum_col <= mr.max_col:
            sum_start_col = mr.min_col
            sum_end_col = mr.max_col
            break

    add_log(logs, f"Exact Date found at {cell_ref(date_row, date_col)}")
    add_log(logs, f"Exact Sum found at {cell_ref(sum_row, sum_col)}")
    add_log(logs, f"Using ONLY Sum section columns: {get_column_letter(sum_start_col)}:{get_column_letter(sum_end_col)}")

    unmerge_and_fill_all(ws)
    selected_dates = set(selected_dates)
    add_log(logs, f"Selected dates: {[d.strftime('%m/%d/%Y') for d in selected_dates]}")

    total_used = 0.0
    active_diesel_col = None
    active_petrol_col = None

    for r in range(sum_row + 1, ws.max_row + 1):
        current_date = normalise_date(ws.cell(r, date_col).value)
        header_pair_found = False

        for c in range(sum_start_col, sum_end_col):
            left_val = ws.cell(r, c).value
            right_val = ws.cell(r, c + 1).value
            left_txt = "" if left_val is None else str(left_val).strip().upper()
            right_txt = "" if right_val is None else str(right_val).strip().upper()

            if left_txt in ("DIESEL", "DIESIL") and right_txt == "PETROL":
                active_diesel_col = c
                active_petrol_col = c + 1
                header_pair_found = True
                add_log(logs, f"SUM header found row {r}: {cell_ref(r, c)}='{left_val}', {cell_ref(r, c + 1)}='{right_val}'")
                break

        if header_pair_found:
            continue
        if current_date not in selected_dates:
            continue
        if active_diesel_col is None or active_petrol_col is None:
            add_log(logs, f"Skipped row {r}: date matched but no Diesel/Petrol header under exact Sum yet.")
            continue

        diesel_raw = ws.cell(r, active_diesel_col).value
        petrol_raw = ws.cell(r, active_petrol_col).value
        diesel_txt = "" if diesel_raw is None else str(diesel_raw).strip().upper()
        petrol_txt = "" if petrol_raw is None else str(petrol_raw).strip().upper()
        if diesel_txt in ("DIESEL", "DIESIL") or petrol_txt == "PETROL":
            continue

        diesel_num = safe_float(diesel_raw)
        petrol_num = safe_float(petrol_raw)
        combined = diesel_num + petrol_num
        total_used += combined
        add_log(logs, f"{org} SUM value row {r}, date {current_date.strftime('%m/%d/%Y')}: Diesel {cell_ref(r, active_diesel_col)}={diesel_raw} -> {diesel_num}, Petrol {cell_ref(r, active_petrol_col)}={petrol_raw} -> {petrol_num}, combined={combined}")

    add_log(logs, f"FINAL USED TOTAL for {org}: {total_used}")
    add_log(logs, "")
    return org, total_used


def extract_fuel_storage(uploaded_file, logs=None):
    org = detect_org(uploaded_file.name)
    add_log(logs, f"========== Fuel Storage | {uploaded_file.name} | detected org: {org} ==========")

    raw = BytesIO(uploaded_file.getvalue())
    wb = load_workbook(raw, data_only=True)
    if "Summary" not in wb.sheetnames:
        raise RuntimeError(f'"Summary" sheet not found in "{uploaded_file.name}".')

    ws = wb["Summary"]
    # Unmerge without filling so the structure stays:
    # Current Fuel In storage (L) | empty | Diesel | Petrol
    for mr in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(mr))

    storage_row, storage_col = find_cell_containing(ws, "Current Fuel In storage")
    if not storage_row:
        raise RuntimeError(f'"Current Fuel In storage (L)" not found in "{uploaded_file.name}".')

    add_log(logs, f"Current Fuel In storage found at {cell_ref(storage_row, storage_col)}")
    diesel_col = None
    for c in range(storage_col + 1, ws.max_column + 1):
        raw_value = ws.cell(storage_row, c).value
        add_log(logs, f"Storage scan {cell_ref(storage_row, c)} raw='{raw_value}'")
        if raw_value is None or str(raw_value).strip() == "":
            continue
        diesel_col = c
        break

    if diesel_col is None:
        add_log(logs, f"No diesel value found for {org}. Storage total = 0")
        return org, 0

    petrol_col = diesel_col + 1
    diesel_raw = ws.cell(storage_row, diesel_col).value
    petrol_raw = ws.cell(storage_row, petrol_col).value if petrol_col <= ws.max_column else None
    diesel = safe_float(diesel_raw)
    petrol = safe_float(petrol_raw)

    add_log(logs, f"Storage values for {org}: Diesel {cell_ref(storage_row, diesel_col)}={diesel_raw} -> {diesel}, Petrol {cell_ref(storage_row, petrol_col)}={petrol_raw} -> {petrol}, combined={diesel + petrol}")
    add_log(logs, f"FINAL STORAGE TOTAL for {org}: {diesel + petrol}")
    add_log(logs, "")
    return org, diesel + petrol


def add_fuel_dashboard_sheet(
    wb,
    latest_uploads,
    selected_dates,
    logs=None,
    sheet_name="Fuel Dashboard",
):
    """Adds the Fuel Dashboard as a new sheet to an existing workbook."""
    all_rows = []
    daily_entries = {}
    fuel_used_by_org = {}
    fuel_storage_by_org = {}

    for uploaded_file in latest_uploads:
        current_rows = extract_current_fuel_from_file(uploaded_file)
        all_rows.extend(current_rows)

        org, org_entries = extract_daily_fuel_entries(uploaded_file, selected_dates)
        daily_entries[org] = org_entries

        org, used_total = extract_fuel_used(uploaded_file, selected_dates, logs=logs)
        fuel_used_by_org[org] = used_total

        org, storage_total = extract_fuel_storage(uploaded_file, logs=logs)
        fuel_storage_by_org[org] = storage_total

    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])

    ws = wb.create_sheet(sheet_name)

    header_fill = PatternFill(fill_type="solid", fgColor="000000")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    # SECTION 1 - Hebrew Fuel Summary Table A1:D4
    headers3 = ["דלק במאגרים", "דלק שנוצל", "דלק שנכנס", ""]
    header_colours = {1: "A3B18A", 2: "F4B183", 3: "A8D5C8", 4: "FFFFFF"}

    for c, h in enumerate(headers3, start=1):
        cell = ws.cell(1, c)
        cell.value = h
        cell.fill = PatternFill(fill_type="solid", fgColor=header_colours[c])
        cell.font = Font(bold=True, size=14)
        cell.border = border
        cell.alignment = center

    total_entered_wfp = sum(daily_entries.get("WFP", {}).values())
    total_entered_unops = sum(daily_entries.get("UNOPS", {}).values())

    table_rows = [
        {
            "org": "UNOPS",
            "entered": total_entered_unops,
            "used": fuel_used_by_org.get("UNOPS", 0),
            "storage": fuel_storage_by_org.get("UNOPS", 0),
            "org_fill": "F47C25",
        },
        {
            "org": "WFP",
            "entered": total_entered_wfp,
            "used": fuel_used_by_org.get("WFP", 0),
            "storage": fuel_storage_by_org.get("WFP", 0),
            "org_fill": "00A9E0",
        },
        {
            "org": 'סה"כ',
            "entered": total_entered_unops + total_entered_wfp,
            "used": fuel_used_by_org.get("UNOPS", 0) + fuel_used_by_org.get("WFP", 0),
            "storage": fuel_storage_by_org.get("UNOPS", 0) + fuel_storage_by_org.get("WFP", 0),
            "org_fill": "FFFFFF",
        },
    ]

    for r, row in enumerate(table_rows, start=2):
        write_number_or_blank(ws.cell(r, 1), row["storage"])
        write_number_or_blank(ws.cell(r, 2), row["used"])
        write_number_or_blank(ws.cell(r, 3), row["entered"])
        org_cell = ws.cell(r, 4)
        org_cell.value = row["org"]
        org_cell.fill = PatternFill(fill_type="solid", fgColor=row["org_fill"])

        for c in range(1, 5):
            cell = ws.cell(r, c)
            cell.border = border
            cell.font = Font(bold=True, size=14)
            cell.alignment = center
            cell.number_format = "#,##0"

    # SECTION 2 - Daily Fuel Entry A7:C...
    daily_start_row = 7
    daily_start_col = 1
    for i, h in enumerate(["Date", "WFP", "UNOPS"]):
        cell = ws.cell(daily_start_row, daily_start_col + i)
        cell.value = h
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center

    row_num = daily_start_row + 1
    for d in selected_dates:
        ws.cell(row_num, daily_start_col).value = d.strftime("%d/%m/%Y")
        write_number_or_blank(ws.cell(row_num, daily_start_col + 1), daily_entries.get("WFP", {}).get(d, 0))
        write_number_or_blank(ws.cell(row_num, daily_start_col + 2), daily_entries.get("UNOPS", {}).get(d, 0))
        for c in range(daily_start_col, daily_start_col + 3):
            cell = ws.cell(row_num, c)
            cell.border = border
            cell.alignment = center
            cell.number_format = "#,##0"
        row_num += 1

    ws.cell(row_num, daily_start_col).value = "Total"
    ws.cell(row_num, daily_start_col + 1).value = f"=SUM(B{daily_start_row + 1}:B{row_num - 1})"
    ws.cell(row_num, daily_start_col + 2).value = f"=SUM(C{daily_start_row + 1}:C{row_num - 1})"
    for c in range(daily_start_col, daily_start_col + 3):
        cell = ws.cell(row_num, c)
        cell.border = border
        cell.font = Font(bold=True)
        cell.alignment = center
        cell.number_format = "#,##0"

    # SECTION 3 - Fuel Station Overview E7:I...
    station_start_row = 7
    station_start_col = 5
    station_headers = ["Organisation", "Fuel Station - Address", "Diesel", "Petrol", "Total Fuel"]
    for i, h in enumerate(station_headers):
        cell = ws.cell(station_start_row, station_start_col + i)
        cell.value = h
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center

    for r_idx, row in enumerate(all_rows, start=station_start_row + 1):
        ws.cell(r_idx, station_start_col).value = row["Organisation"]
        ws.cell(r_idx, station_start_col + 1).value = row["Fuel Station - Address"]
        write_number_or_blank(ws.cell(r_idx, station_start_col + 2), row["Diesel"])
        write_number_or_blank(ws.cell(r_idx, station_start_col + 3), row["Petrol"])
        write_number_or_blank(ws.cell(r_idx, station_start_col + 4), row["Total Fuel"])

        for c in range(station_start_col, station_start_col + 5):
            cell = ws.cell(r_idx, c)
            cell.border = border
            cell.alignment = center
            cell.number_format = "#,##0"

    widths = {"A": 18, "B": 18, "C": 18, "D": 18, "E": 18, "F": 32, "G": 15, "H": 15, "I": 15}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for r in range(1, ws.max_row + 1):
        ws.row_dimensions[r].height = 24
    for r in range(1, 5):
        ws.row_dimensions[r].height = 35
    ws.freeze_panes = "A7"

    return wb

# ============================================================
# Settings
# ============================================================
STATUS_COL_NAME = "סטטוס"
UNKNOWN_ENTITIES_STATUS = "ארגונים/גופים לא מוכרים (INGOs/חברות מקומיות)"
DISTRIBUTION_SHEET_NAME = "Distribution Summary"
FUZZY_THRESHOLD = 0.78

APPROVAL_SHEETS = {
    "UNOPS": ["UNOPS Total Distribution", "UNOPS Total Distribution "],
    "WFP": ["WFP Total Distribution"],
}

# The original cleaner may prefix local/non-standard agencies like:
#   NNGOs - Culture and Free Thought Association (CFTA)
# while the approved list stores:
#   Culture and Free Thought Association (CFTA)
# More prefixes are added dynamically from the approval workbook clusters/interventions.
BASE_GENERATED_PREFIXES = {
    "NNGOS",
    "NNGO",
    "FINANCIAL INSTITUTIONS",
    "FINANCIAL INSTITUTION",
    "UN-OHCHR",
    "UN OHCHR",
    "UN - SISTERS LOGISTICS",
    "UN SISTERS LOGISTICS",
    "UN - SISTERS AGENCIES",
    "UN SISTERS AGENCIES",
    "PROTECTION",
    "SHELTER",
    "EDUCATION",
    "NUTRITION",
    "LOGISTICS",
    "FOOD SECURITY",
    "LOCAL AUTHORITIES",
    "CCCM",
}

STATUS_FILL_RULES = [
    ("מסורב", "F4B084"),
    ("מוקפא", "B4C6E7"),
    ("מאפיות", "FFE699"),
    ("מאושר", "C6E0B4"),
    ("שותפים", "E2F0D9"),
    ("בריאות", "C6E0B4"),
    ("תקשורת", "DDEBF7"),
    ("סניטציה", "FCE4D6"),
    ("סוכנויות", "D9EAD3"),
    ("ארגונים/גופים לא מוכרים", "C044E8"),
    ("חברות מקומיות", "C044E8"),
    ("לא מוכרים", "C044E8"),
    ("NOT FOUND", "D9D9D9"),
    ("UNKNOWN", "D9D9D9"),
]


# ============================================================
# Normalisation + matching helpers
# ============================================================
def _norm_prefix(value) -> str:
    if value is None:
        return ""
    s = str(value).strip().upper()
    s = s.replace("’", "'").replace("`", "'")
    s = s.replace("‐", "-").replace("‑", "-").replace("–", "-").replace("—", "-")
    s = re.sub(r"[^A-Z0-9א-ת]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def strip_generated_prefix(value, prefixes=None) -> str:
    """Remove prefixes added by the cleaner, for example 'NNGOs - ...'."""
    if value is None:
        return ""

    s = str(value).strip()
    if not s:
        return ""

    prefixes = prefixes or BASE_GENERATED_PREFIXES
    norm_prefixes = sorted({_norm_prefix(p) for p in prefixes if p}, key=len, reverse=True)

    changed = True
    while changed:
        changed = False
        s_norm = _norm_prefix(s)

        for prefix in norm_prefixes:
            if not prefix:
                continue

            # Prefix may contain punctuation in the original text, so compare a normalised left side.
            # We only remove when the actual text has a dash/colon separator after the prefix area.
            parts = re.split(r"\s*[-–—:]\s*", s, maxsplit=1)
            if len(parts) == 2 and _norm_prefix(parts[0]) == prefix:
                s = parts[1].strip()
                changed = True
                break

            # Handles cases like 'UN - Sisters Logistics - UN - Sisters Agencies'.
            # Try removing a normalised prefix from the beginning when the next visible char is a separator.
            pattern = re.compile(r"^\s*" + re.escape(str(prefix)) + r"\s*[-–—:]\s*", re.IGNORECASE)
            new_s = pattern.sub("", s).strip()
            if new_s != s:
                s = new_s
                changed = True
                break

    return s


def norm_text(value, prefixes=None) -> str:
    """Normalise names while ignoring cleaner-added category prefixes."""
    if value is None:
        return ""

    s = strip_generated_prefix(value, prefixes=prefixes).upper()
    s = s.replace("’", "'").replace("`", "'")
    s = s.replace("‐", "-").replace("‑", "-").replace("–", "-").replace("—", "-")
    s = s.replace("&", " AND ")

    # Usually acronyms in brackets differ between files; remove them to match the full name.
    s = re.sub(r"\([^)]*\)", " ", s)
    s = re.sub(r"[^A-Z0-9א-ת]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()

    replacements = {
        "INTL": "INTERNATIONAL",
        "INT": "INTERNATIONAL",
        "ORG": "ORGANIZATION",
        "ASSOC": "ASSOCIATION",
        "PROGRAMME": "PROGRAM",
        "SOCAITY": "SOCIETY",
        "CLOBAL": "GLOBAL",
        "PH": "PROJECT HOPE",
        "ACF": "ACTION AGAINST HUNGER",
        "SI": "SOLIDARITES INTERNATIONALE",
        "GC": "GLOBAL COMMUNITIES",
        "MG": "MED GLOBAL",
        "PIB": "PALESTINE ISLAMIC BANK",
    }
    words = [replacements.get(w, w) for w in s.split()]
    return " ".join(words)


def compact_key(*parts, prefixes=None) -> str:
    return " | ".join(norm_text(p, prefixes=prefixes) for p in parts if norm_text(p, prefixes=prefixes))


def clean_header(value) -> str:
    return norm_text(value).replace(" ", "")


def best_fuzzy_match(query_key: str, choices: list[str]):
    if not query_key or not choices:
        return None, 0.0

    shortlist = get_close_matches(query_key, choices, n=8, cutoff=0.55)
    if not shortlist:
        shortlist = choices

    best_key = None
    best_score = 0.0
    for choice in shortlist:
        score = SequenceMatcher(None, query_key, choice).ratio()
        if score > best_score:
            best_score = score
            best_key = choice

    if best_score >= FUZZY_THRESHOLD:
        return best_key, best_score
    return None, best_score


# ============================================================
# Workbook helpers
# ============================================================
def get_sheet_by_names(wb, possible_names):
    name_map = {name.strip().upper(): name for name in wb.sheetnames}
    for wanted in possible_names:
        real = name_map.get(wanted.strip().upper())
        if real:
            return wb[real]
    return None


def find_header_row_and_columns(ws, wanted_headers, max_scan_rows=20):
    wanted_clean = {canonical: [clean_header(x) for x in aliases] for canonical, aliases in wanted_headers.items()}

    for r in range(1, min(ws.max_row, max_scan_rows) + 1):
        row_map = {}
        for c in range(1, ws.max_column + 1):
            cell_clean = clean_header(ws.cell(r, c).value)
            if not cell_clean:
                continue
            for canonical, aliases in wanted_clean.items():
                if cell_clean in aliases:
                    row_map[canonical] = c

        required = [k for k in wanted_headers if k != "Status"]
        if all(k in row_map for k in required):
            return r, row_map

    return None, {}


def copy_cell_style(src, dst):
    if src is None or dst is None:
        return
    dst.font = copy(src.font)
    dst.fill = copy(src.fill)
    dst.border = copy(src.border)
    dst.alignment = copy(src.alignment)
    dst.number_format = src.number_format
    dst.protection = copy(src.protection)


def safe_float(value):
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        s = value.strip().replace(",", "")
        if not s:
            return 0.0
        try:
            return float(s)
        except Exception:
            return 0.0
    return 0.0


def status_fill(status):
    txt = "" if status is None else str(status).strip().upper()
    for needle, colour in STATUS_FILL_RULES:
        if needle.upper() in txt:
            return PatternFill(fill_type="solid", fgColor=colour)
    return PatternFill(fill_type="solid", fgColor="D9D9D9")


def _shift_merge_range_for_insert(range_string: str, insert_col: int) -> str:
    """Return the same merge range after a column insertion, preserving existing merges."""
    min_col, min_row, max_col, max_row = range_boundaries(range_string)

    # If the full merged range is to the right of the inserted column, move it right.
    if min_col >= insert_col:
        min_col += 1
        max_col += 1
    # If a merged range crosses the insertion point, expand it by one column.
    elif min_col < insert_col <= max_col:
        max_col += 1

    return (
        f"{get_column_letter(min_col)}{min_row}:"
        f"{get_column_letter(max_col)}{max_row}"
    )


def _insert_column_preserve_merges(ws, insert_col: int):
    """openpyxl does not reliably shift merged ranges on insert_cols; do it manually."""
    merge_ranges = [str(mr) for mr in ws.merged_cells.ranges]
    for rng in merge_ranges:
        ws.unmerge_cells(rng)

    ws.insert_cols(insert_col, 1)

    for rng in merge_ranges:
        shifted = _shift_merge_range_for_insert(rng, insert_col)
        try:
            ws.merge_cells(shifted)
        except Exception:
            # If a corrupted/overlapping merge exists in an input file, skip only that merge.
            pass


def _last_text_header_col(ws, header_row, exclude_status=True):
    """Return the last header column that contains visible text.

    This is intentionally based on the header row, not ws.max_column, because
    generated Fuels Summary workbooks can contain charts/empty formatted columns
    to the right.
    """
    last_col = 1
    for c in range(1, ws.max_column + 1):
        value = ws.cell(header_row, c).value
        if value is None or str(value).strip() == "":
            continue
        if exclude_status and str(value).strip() == STATUS_COL_NAME:
            continue
        last_col = max(last_col, c)
    return last_col


def get_or_insert_status_column(ws, header_row, description_col=None):
    """Add סטטוס AFTER the last real header column.

    Earlier versions inserted סטטוס next to DESCRIPTION. In this workbook that
    shifts the Unified Fuel / Total Sum Per Category area and can make Unified
    Fuel appear merged by cluster. Appending סטטוס after the last text header
    keeps all calculation columns exactly where they are.
    """
    existing_status_cols = [
        c for c in range(1, ws.max_column + 1)
        if str(ws.cell(header_row, c).value).strip() == STATUS_COL_NAME
    ]

    last_text_col = _last_text_header_col(ws, header_row, exclude_status=True)
    desired_status_col = last_text_col + 1

    # If the workbook was already created with this fixed version, reuse the column.
    for c in existing_status_cols:
        if c == desired_status_col or c > last_text_col:
            return c, False

    # If a previous/bad version inserted סטטוס inside the calculation columns, stop
    # rather than producing a misleading dashboard from already-shifted/merged data.
    if existing_status_cols:
        raise RuntimeError(
            "This workbook already contains a סטטוס column inside the calculation area. "
            "Please upload the original Fuels Summary workbook and rerun this fixed version."
        )

    status_col = desired_status_col
    _insert_column_preserve_merges(ws, status_col)

    status_header = ws.cell(header_row, status_col)
    status_header.value = STATUS_COL_NAME
    style_source_col = description_col if description_col else last_text_col
    copy_cell_style(ws.cell(header_row, style_source_col), status_header)
    status_header.alignment = Alignment(horizontal="center", vertical="center")
    ws.column_dimensions[get_column_letter(status_col)].width = 22
    return status_col, True


# ============================================================
# Approval index
# ============================================================
def read_approval_rows(approval_file):
    raw = BytesIO(approval_file.getvalue())
    wb = load_workbook(raw, data_only=True)

    rows = []
    prefixes = set(BASE_GENERATED_PREFIXES)

    # ---------- UNOPS approval sheet ----------
    ws_unops = get_sheet_by_names(wb, APPROVAL_SHEETS["UNOPS"])
    if ws_unops is not None:
        headers = {
            "Cluster": ["Cluster"],
            "Agency": ["AGENCY", "Agency"],
            "Description": ["DESCRIPTION", "Description"],
            "Status": [STATUS_COL_NAME, "Status"],
        }
        header_row, cols = find_header_row_and_columns(ws_unops, headers)
        if header_row:
            for r in range(header_row + 1, ws_unops.max_row + 1):
                cluster = ws_unops.cell(r, cols.get("Cluster", 0)).value if cols.get("Cluster") else None
                agency = ws_unops.cell(r, cols.get("Agency", 0)).value if cols.get("Agency") else None
                desc = ws_unops.cell(r, cols.get("Description", 0)).value if cols.get("Description") else None
                status = ws_unops.cell(r, cols.get("Status", 0)).value if cols.get("Status") else None
                if not any([cluster, agency, desc, status]):
                    continue

                prefixes.add(str(cluster).strip())
                rows.append({
                    "Source Org": "UNOPS",
                    "Approved Cluster/Intervention": cluster,
                    "Approved Agency": agency,
                    "Approved Description": desc,
                    "Review Status": status or "Unknown",
                    "Approval Sheet Row": r,
                })

    # ---------- WFP approval sheet ----------
    ws_wfp = get_sheet_by_names(wb, APPROVAL_SHEETS["WFP"])
    if ws_wfp is not None:
        headers = {
            "Intervention": ["INTERVENTION", "Intervention"],
            "Description": ["DESCRIPTION", "Description"],
            "Status": [STATUS_COL_NAME, "Status"],
        }
        header_row, cols = find_header_row_and_columns(ws_wfp, headers)
        if header_row:
            for r in range(header_row + 1, ws_wfp.max_row + 1):
                intervention = ws_wfp.cell(r, cols.get("Intervention", 0)).value if cols.get("Intervention") else None
                desc = ws_wfp.cell(r, cols.get("Description", 0)).value if cols.get("Description") else None
                status = ws_wfp.cell(r, cols.get("Status", 0)).value if cols.get("Status") else None
                if not any([intervention, desc, status]):
                    continue

                prefixes.add(str(intervention).strip())
                rows.append({
                    "Source Org": "WFP",
                    "Approved Cluster/Intervention": intervention,
                    "Approved Agency": None,
                    "Approved Description": desc,
                    "Review Status": status or "Unknown",
                    "Approval Sheet Row": r,
                })

    if not rows:
        raise RuntimeError("No approval rows were found in the approval workbook.")

    return rows, prefixes


def build_approval_index(approval_file):
    approval_rows, prefixes = read_approval_rows(approval_file)

    index = {
        "UNOPS": {"exact": {}, "choices": []},
        "WFP": {"exact": {}, "choices": []},
    }

    for row in approval_rows:
        org = row["Source Org"]
        cluster = row["Approved Cluster/Intervention"]
        agency = row["Approved Agency"]
        desc = row["Approved Description"]

        if org == "UNOPS":
            keys = {
                compact_key(agency, desc, prefixes=prefixes),
                compact_key(agency, prefixes=prefixes),
                compact_key(desc, prefixes=prefixes),
                compact_key(cluster, agency, prefixes=prefixes),
                compact_key(cluster, agency, desc, prefixes=prefixes),
            }
        else:
            keys = {
                compact_key(cluster, desc, prefixes=prefixes),
                compact_key(desc, prefixes=prefixes),
                compact_key(cluster, prefixes=prefixes),
            }

        for key in keys:
            if key:
                index[org]["exact"].setdefault(key, row)

    for org in ("UNOPS", "WFP"):
        index[org]["choices"] = list(index[org]["exact"].keys())

    return index, prefixes, pd.DataFrame(approval_rows)


# ============================================================
# Distribution Summary comparison
# ============================================================
def detect_row_source_org(cluster_value) -> str:
    # In the final Fuels Summary workbook, WFP rows have Cluster = WFP.
    return "WFP" if norm_text(cluster_value) == "WFP" else "UNOPS"


def review_summary_row(cluster, agency, desc, approval_index, prefixes):
    # Final Fuels Summary can contain internal UN-agency operational rows under
    # Cluster = UN Agencies. These should be reviewed as approved UN agency fuel
    # instead of being matched as a regular NGO/organisation row.
    agency_compact = _compact_alpha_num(agency)
    if norm_text(cluster) == "UN AGENCIES" and (
        norm_text(agency) == "WFP" or "OHCHR" in agency_compact
    ):
        return {
            "Detected Org": "UNOPS",
            "Review Status": 'סוכנויות או"ם מאושרות',
            "Match Type": "Rule",
            "Match Score": 1.0,
            "Matched Cluster/Intervention": cluster,
            "Matched Agency": agency,
            "Matched Description": desc,
            "Approval Sheet Row": None,
        }

    org = detect_row_source_org(cluster)
    exact = approval_index[org]["exact"]
    choices = approval_index[org]["choices"]

    cleaned_agency = strip_generated_prefix(agency, prefixes=prefixes)

    if org == "WFP":
        candidate_keys = [
            compact_key(agency, desc, prefixes=prefixes),
            compact_key(cleaned_agency, desc, prefixes=prefixes),
            compact_key(desc, prefixes=prefixes),
            compact_key(cleaned_agency, prefixes=prefixes),
        ]
    else:
        candidate_keys = [
            compact_key(agency, desc, prefixes=prefixes),
            compact_key(cleaned_agency, desc, prefixes=prefixes),
            compact_key(cleaned_agency, prefixes=prefixes),
            compact_key(desc, prefixes=prefixes),
            compact_key(cluster, cleaned_agency, prefixes=prefixes),
            compact_key(cluster, cleaned_agency, desc, prefixes=prefixes),
        ]

    for key in candidate_keys:
        if key and key in exact:
            match = exact[key]
            return {
                "Detected Org": org,
                "Review Status": match["Review Status"],
                "Match Type": "Exact",
                "Match Score": 1.0,
                "Matched Cluster/Intervention": match["Approved Cluster/Intervention"],
                "Matched Agency": match["Approved Agency"],
                "Matched Description": match["Approved Description"],
                "Approval Sheet Row": match["Approval Sheet Row"],
            }

    # Fuzzy match using all candidate keys. This handles small typos and non-exact formats.
    best_key = None
    best_score = 0.0
    for query_key in candidate_keys:
        fuzzy_key, score = best_fuzzy_match(query_key, choices)
        if score > best_score:
            best_score = score
            best_key = fuzzy_key

    if best_key:
        match = exact[best_key]
        return {
            "Detected Org": org,
            "Review Status": match["Review Status"],
            "Match Type": "Fuzzy",
            "Match Score": round(best_score, 3),
            "Matched Cluster/Intervention": match["Approved Cluster/Intervention"],
            "Matched Agency": match["Approved Agency"],
            "Matched Description": match["Approved Description"],
            "Approval Sheet Row": match["Approval Sheet Row"],
        }

    no_match_status = UNKNOWN_ENTITIES_STATUS if norm_text(cluster) == "INGOS" else "Not Found"

    return {
        "Detected Org": org,
        "Review Status": no_match_status,
        "Match Type": "No Match",
        "Match Score": round(best_score, 3),
        "Matched Cluster/Intervention": None,
        "Matched Agency": None,
        "Matched Description": None,
        "Approval Sheet Row": None,
    }




# ============================================================
# Final workbook formatting helpers
# ============================================================
def _has_hebrew(value) -> bool:
    if value is None:
        return False
    return any("\u0590" <= ch <= "\u05FF" for ch in str(value))


def _is_number_like(value) -> bool:
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return True
    txt = str(value).strip().replace(",", "").replace("%", "")
    if txt == "":
        return False
    try:
        float(txt)
        return True
    except Exception:
        return False


def _font_with(cell, name=None, size=None, color=None, bold=None, italic=None):
    f = cell.font or Font()
    return Font(
        name=name if name is not None else f.name,
        sz=size if size is not None else f.sz,
        b=bold if bold is not None else f.bold,
        i=italic if italic is not None else f.italic,
        u=f.underline,
        strike=f.strike,
        color=color if color is not None else f.color,
        vertAlign=f.vertAlign,
        outline=f.outline,
        shadow=f.shadow,
        condense=f.condense,
        extend=f.extend,
        charset=f.charset,
        family=f.family,
        scheme=f.scheme,
    )


def _thin_black_border():
    side = Side(style="thin", color="000000")
    return Border(left=side, right=side, top=side, bottom=side)


def _safe_unmerge_intersecting_columns(ws, cols_to_delete):
    cols = set(cols_to_delete)
    for mr in list(ws.merged_cells.ranges):
        if any(mr.min_col <= c <= mr.max_col for c in cols):
            ws.unmerge_cells(str(mr))


def _delete_columns_by_headers(ws, header_row, header_names):
    """Delete columns whose header text exactly matches any value in header_names."""
    names = {str(x).strip().upper() for x in header_names}
    cols = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if v is not None and str(v).strip().upper() in names:
            cols.append(c)
    if not cols:
        return
    _safe_unmerge_intersecting_columns(ws, cols)
    for c in sorted(cols, reverse=True):
        ws.delete_cols(c, 1)


def _last_nonempty_header_col(ws, header_row=1):
    last = 1
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if v is not None and str(v).strip() != "":
            last = c
    return last


def _set_autofilter_to_real_headers(ws, header_row=1):
    last = _last_nonempty_header_col(ws, header_row)
    if last >= 1 and ws.max_row >= header_row:
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(last)}{ws.max_row}"
    else:
        ws.auto_filter.ref = None


def _find_header_col_exact(ws, header_text, header_row=1):
    target = str(header_text).strip().upper()
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if v is not None and str(v).strip().upper() == target:
            return c
    return None


def _auto_fit_columns(ws, min_width=10, max_width=55):
    for c in range(1, ws.max_column + 1):
        letter = get_column_letter(c)
        max_len = 0
        for r in range(1, ws.max_row + 1):
            value = ws.cell(r, c).value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        if max_len:
            ws.column_dimensions[letter].width = min(max(max_len + 3, min_width), max_width)


def _row_has_any_real_value(ws, row_num, max_col=None):
    max_col = max_col or ws.max_column
    for c in range(1, max_col + 1):
        value = ws.cell(row_num, c).value
        if value is not None and str(value).strip() != "":
            return True
    return False


def remerge_total_sum_per_category(ws, header_row=1):
    """Rebuild the Total Sum Per Category merges without touching other columns.

    This is run at the very end because column deletion/style operations can
    disturb merged ranges. The values are recalculated from row-level Unified
    Fuel, then Total Sum Per Category is merged only by consecutive Cluster
    groups.
    """
    total_col = _find_header_col_exact(ws, "Total Sum Per Category", header_row=header_row)
    if total_col is None:
        return

    cluster_col = _find_header_col_exact(ws, "Cluster", header_row=header_row)
    if cluster_col is None:
        cluster_col = _find_header_col_exact(ws, "INTERVENTION", header_row=header_row)
    if cluster_col is None:
        cluster_col = 1

    unified_col = _find_header_col_exact(ws, "Unified Fuel", header_row=header_row)

    # Unmerge only ranges intersecting Total Sum Per Category, preserving values/styles.
    for mr in list(ws.merged_cells.ranges):
        if mr.min_col <= total_col <= mr.max_col and mr.max_row > header_row:
            top = ws.cell(mr.min_row, total_col)
            top_value = top.value
            top_style = copy(top._style)
            top_font = copy(top.font)
            top_alignment = copy(top.alignment)
            top_number_format = top.number_format
            top_protection = copy(top.protection)
            top_fill = copy(top.fill)
            ws.unmerge_cells(str(mr))
            for r in range(mr.min_row, mr.max_row + 1):
                cell = _ensure_real_cell(ws, r, total_col)
                cell.value = top_value
                cell._style = copy(top_style)
                cell.font = copy(top_font)
                cell.alignment = copy(top_alignment)
                cell.number_format = top_number_format
                cell.protection = copy(top_protection)
                cell.fill = copy(top_fill)

    # Recalculate total per cluster from the reliable row-level Unified Fuel column.
    if unified_col is not None:
        totals = {}
        for r in range(header_row + 1, ws.max_row + 1):
            if not _row_has_any_real_value(ws, r, max_col=ws.max_column):
                continue
            key = "" if ws.cell(r, cluster_col).value is None else str(ws.cell(r, cluster_col).value).strip()
            totals[key] = totals.get(key, 0.0) + safe_float(ws.cell(r, unified_col).value)

        for r in range(header_row + 1, ws.max_row + 1):
            key = "" if ws.cell(r, cluster_col).value is None else str(ws.cell(r, cluster_col).value).strip()
            if key in totals:
                target = _ensure_real_cell(ws, r, total_col)
                target.value = totals[key]

    # Rebuild merges by consecutive cluster groups.
    r = header_row + 1
    while r <= ws.max_row:
        if not _row_has_any_real_value(ws, r, max_col=ws.max_column):
            r += 1
            continue

        key = "" if ws.cell(r, cluster_col).value is None else str(ws.cell(r, cluster_col).value).strip().upper()
        start = r
        end = r
        rr = r + 1
        while rr <= ws.max_row:
            if not _row_has_any_real_value(ws, rr, max_col=ws.max_column):
                break
            next_key = "" if ws.cell(rr, cluster_col).value is None else str(ws.cell(rr, cluster_col).value).strip().upper()
            if next_key != key:
                break
            end = rr
            rr += 1

        top = ws.cell(start, total_col)
        top_value = top.value
        top_style = copy(top._style)
        top_font = copy(top.font)
        top_alignment = copy(top.alignment)
        top_number_format = top.number_format
        top_protection = copy(top.protection)
        top_fill = copy(top.fill)

        if end > start:
            ws.merge_cells(start_row=start, start_column=total_col, end_row=end, end_column=total_col)

        top = ws.cell(start, total_col)
        top.value = top_value
        top._style = copy(top_style)
        top.font = copy(top_font)
        top.alignment = copy(top_alignment)
        top.number_format = top_number_format
        top.protection = copy(top_protection)
        top.fill = copy(top_fill)

        r = end + 1


def style_distribution_summary_final(ws):
    """Final requested styling for Distribution Summary."""
    # Reapply the merge first, then style; no borders are applied to the
    # Total Sum Per Category column, but all other requested styling is applied.
    remerge_total_sum_per_category(ws, header_row=1)

    border = _thin_black_border()
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    total_cat_col = _find_header_col_exact(ws, "Total Sum Per Category", header_row=1)

    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue

            # Style only cells that actually contain values/text; do not format the full sheet.
            if cell.value is None or str(cell.value).strip() == "":
                continue

            cell.alignment = center
            font_color = "FFFFFF" if cell.row == 1 else "000000"
            cell.font = _font_with(cell, name="Calibri", size=11, color=font_color, bold=False if cell.row != 1 else True, italic=False)

            # Apply all borders only outside Total Sum Per Category.
            if total_cat_col is None or cell.column != total_cat_col:
                cell.border = border

    _set_autofilter_to_real_headers(ws, header_row=1)

def style_sector_summary_final(ws):
    """Final requested styling for Sector Summary."""
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    un_agencies_fill = PatternFill(fill_type="solid", fgColor="013D59")

    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            cell.alignment = center

            color = "FFFFFF" if cell.row == 1 else "000000"
            font_name = "Aduma" if (_has_hebrew(cell.value) or _is_number_like(cell.value) or (isinstance(cell.value, str) and cell.value.startswith("="))) else None
            cell.font = _font_with(cell, name=font_name, size=20, color=color, bold=False, italic=False)

    # Header row text should be white.
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(1, c)
        if cell.value is not None and str(cell.value).strip() != "":
            cell.font = _font_with(cell, color="FFFFFF", bold=False, italic=False)

    # סוכנויות או"ם row: dark blue fill #013D59 and white text/value.
    for r in range(1, ws.max_row + 1):
        val = ws.cell(r, 1).value
        if val is not None and str(val).strip() == 'סוכנויות או"ם':
            for c in range(1, min(ws.max_column, 3) + 1):
                cell = ws.cell(r, c)
                cell.fill = un_agencies_fill
                cell.font = _font_with(cell, color="FFFFFF", bold=False, italic=False)

    _auto_fit_columns(ws, min_width=12, max_width=60)


def style_fuel_dashboard_final(ws):
    """Final requested styling for Fuel Dashboard."""
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Top table A1:D4 -> Aduma 16 for Hebrew/numbers, no bold.
    for r in range(1, min(ws.max_row, 4) + 1):
        for c in range(1, min(ws.max_column, 4) + 1):
            cell = ws.cell(r, c)
            cell.alignment = center
            if _has_hebrew(cell.value) or _is_number_like(cell.value):
                cell.font = _font_with(cell, name="Aduma", size=16, bold=False, italic=False)
            else:
                cell.font = _font_with(cell, size=16, bold=False, italic=False)

    # Bottom tables from row 7 down -> size 12, Aduma for Hebrew/numbers.
    for r in range(7, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            cell.alignment = center
            if _has_hebrew(cell.value) or _is_number_like(cell.value):
                cell.font = _font_with(cell, name="Aduma", size=12, bold=False, italic=False)
            elif cell.value is not None:
                cell.font = _font_with(cell, size=12, bold=False, italic=False)


def style_status_dashboard_final(ws):
    """Final requested styling for סטטוס לפי אישורים."""
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = center
            if _has_hebrew(cell.value) or _is_number_like(cell.value):
                cell.font = _font_with(cell, name="Aduma", size=12, bold=False, italic=False)
            elif cell.value is not None:
                cell.font = _font_with(cell, size=12, bold=False, italic=False)

def move_sheet_before(wb, sheet_name, before_sheet_name):
    if sheet_name not in wb.sheetnames or before_sheet_name not in wb.sheetnames:
        return
    sheet = wb[sheet_name]
    wb._sheets.remove(sheet)
    before_idx = wb.sheetnames.index(before_sheet_name)
    wb._sheets.insert(before_idx, sheet)


def apply_final_workbook_layout(wb):
    if "Distribution Summary" in wb.sheetnames:
        style_distribution_summary_final(wb["Distribution Summary"])
    if "Sector Summary" in wb.sheetnames:
        style_sector_summary_final(wb["Sector Summary"])
    if "Fuel Dashboard" in wb.sheetnames:
        style_fuel_dashboard_final(wb["Fuel Dashboard"])
        move_sheet_before(wb, "Fuel Dashboard", "Sector Summary")
    if "סטטוס לפי אישורים" in wb.sheetnames:
        style_status_dashboard_final(wb["סטטוס לפי אישורים"])

# ============================================================
# Status Dashboard sheet
# ============================================================
def _text_contains(value, *needles) -> bool:
    txt = "" if value is None else str(value).strip().upper()
    return any(str(n).strip().upper() in txt for n in needles if n)


def _normalised_cluster(value) -> str:
    return "" if value is None else str(value).strip().upper()


def _find_required_dashboard_cols(ws, header_row=1):
    cols = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(header_row, c).value
        if h is None:
            continue
        h_txt = str(h).strip().upper()
        if h_txt == "CLUSTER":
            cols["Cluster"] = c
        elif h_txt == "AGENCY":
            cols["Agency"] = c
        elif h_txt == "DESCRIPTION":
            cols["Description"] = c
        elif h_txt == STATUS_COL_NAME:
            cols["Status"] = c
        elif h_txt == "UNIFIED FUEL":
            cols["Unified Fuel"] = c
    missing = [k for k in ["Cluster", "Status", "Unified Fuel"] if k not in cols]
    if missing:
        raise RuntimeError(f"Cannot build dashboard. Missing columns: {', '.join(missing)}")
    return cols


def _unified_fuel_column_has_bad_merges(ws, fuel_col, header_row=1):
    """Detect the old bug where Unified Fuel was merged by cluster after inserting סטטוס."""
    bad = []
    for mr in ws.merged_cells.ranges:
        if mr.min_col <= fuel_col <= mr.max_col and mr.max_row > header_row:
            bad.append(str(mr))
    return bad


def _dashboard_category(cluster, agency, status):
    """Map each reviewed row into the requested Hebrew dashboard buckets."""
    c = _normalised_cluster(cluster)
    a = norm_text(agency)
    s = "" if status is None else str(status).strip()
    su = s.upper()

    # Special rule: operational rows under UN Agencies for WFP/OHCHR are approved
    # UN agency fuel, not routine/partner fuel and not unknown entities.
    if c == "UN AGENCIES" and (a == "WFP" or "OHCHR" in _compact_alpha_num(agency)):
        return "un_approved"

    # Dedicated sector/status buckets first.
    if "תקשורת" in s:
        return "communications"
    if "בריאות" in s:
        return "health"
    if "סניטציה" in s:
        return "wash_unassigned"
    if "מאפיות" in s:
        return "wfp_bakeries"

    # WFP-specific grouping.
    if c == "WFP":
        if "פעילות" in s:
            return "wfp_routine"
        if "שותפים" in s or "ארגון" in s:
            return "wfp_partners"

    # UN buckets.
    if "סוכנויות" in s and ("לא מפורט" in s or "לא מוכר" in s or "UN SISTERS" in su):
        return "un_unrecognized"
    if "סוכנויות" in s and "מאושרות" in s:
        return "un_approved"

    # Unknown/non-listed organisations.
    if "לא מוכרים" in s or "לא מוכרות" in s or "חברות מקומיות" in s or "NOT FOUND" in su:
        return "unknown_entities"

    # NGO approval review buckets. If a status contains both מסורב and מוקפא
    # (for example TDH), count it as מסורב.
    if "מסורב" in s:
        return "ngo_disapproved"
    if "מוקפא" in s:
        return "ngo_frozen"
    if "מאושר" in s:
        return "ngo_approved"

    return "unknown_entities"


def _write_dashboard_row(ws, row, sector, detail, amount, total, fill_hex, font_hex="FFFFFF", merge_detail=False):
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    fill = PatternFill(fill_type="solid", fgColor=fill_hex)

    if merge_detail:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        c = ws.cell(row, 1)
        c.value = detail
        for col in (1, 2):
            cell = ws.cell(row, col)
            cell.fill = fill
            cell.font = Font(color=font_hex, bold=True, size=12)
            cell.alignment = center
            cell.border = border
    else:
        ws.cell(row, 1).value = sector
        ws.cell(row, 2).value = detail
        for col in (1, 2):
            cell = ws.cell(row, col)
            cell.fill = fill
            cell.font = Font(color=font_hex, bold=True, size=12)
            cell.alignment = center
            cell.border = border

    qty_cell = ws.cell(row, 3)
    pct_cell = ws.cell(row, 4)
    qty_cell.value = amount if amount else None
    pct_cell.value = (amount / total) if total else None

    for cell in (qty_cell, pct_cell):
        cell.fill = fill if cell.column == 3 else PatternFill(fill_type="solid", fgColor="D9D9D9")
        cell.font = Font(color=(font_hex if cell.column == 3 else "000000"), bold=True, size=12)
        cell.alignment = center
        cell.border = border
    qty_cell.number_format = "#,##0"
    pct_cell.number_format = "0.0%"


def add_status_dashboard_sheet(wb, source_ws, header_row=1, sheet_name="סטטוס לפי אישורים"):
    """
    Create the requested summary sheet using row-level Unified Fuel values.
    The function intentionally sums Unified Fuel per reviewed row, NOT Total Sum Per Category.
    """
    cols = _find_required_dashboard_cols(source_ws, header_row=header_row)
    fuel_col = cols["Unified Fuel"]
    bad_merges = _unified_fuel_column_has_bad_merges(source_ws, fuel_col, header_row=header_row)
    if bad_merges:
        raise RuntimeError(
            "Unified Fuel is merged in the reviewed file, so row-level dashboard values cannot be trusted. "
            "This version fixes the merge issue for new outputs. Please rerun it on the original Fuels Summary file. "
            f"Problematic ranges: {', '.join(bad_merges[:8])}"
        )

    sums = {
        "ngo_approved": 0.0,
        "ngo_disapproved": 0.0,
        "ngo_frozen": 0.0,
        "wash_unassigned": 0.0,
        "wfp_routine": 0.0,
        "wfp_partners": 0.0,
        "wfp_bakeries": 0.0,
        "un_approved": 0.0,
        "un_unrecognized": 0.0,
        "health": 0.0,
        "unknown_entities": 0.0,
        "communications": 0.0,
        "unclassified": 0.0,
    }

    source_rows = []
    for r in range(header_row + 1, source_ws.max_row + 1):
        cluster = source_ws.cell(r, cols["Cluster"]).value
        agency = source_ws.cell(r, cols.get("Agency")).value if cols.get("Agency") else None
        status = source_ws.cell(r, cols["Status"]).value
        fuel = safe_float(source_ws.cell(r, fuel_col).value)
        if fuel == 0 and not status:
            continue
        category = _dashboard_category(cluster, agency, status)
        sums[category] = sums.get(category, 0.0) + fuel
        source_rows.append({
            "Workbook Row": r,
            "Cluster": cluster,
            "Agency": agency,
            "Status": status,
            "Unified Fuel": fuel,
            "Dashboard Category": category,
        })

    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.rightToLeft = False

    # Main table structure matching the requested screenshot.
    headers = ["סקטור", "פירוט", "כמות", "אחוז"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(1, c)
        cell.value = h
        cell.fill = PatternFill(fill_type="solid", fgColor="262626")
        cell.font = Font(color="FFFFFF", bold=True, size=13)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000"),
        )

    row_specs = [
        ("NGO", "ארגונים מאושרים", "ngo_approved", "2B8C65", "FFFFFF", False),
        ("NGO", "ארגונים מסורבים", "ngo_disapproved", "FF0000", "FFFFFF", False),
        ("NGO", "ארגונים מוקפאים", "ngo_frozen", "FF0000", "FFFFFF", False),
        (None, "סניטציה ללא שיוך ארגוני", "wash_unassigned", "7563B8", "FFFFFF", True),
        ("WFP", "פעילות שוטפת", "wfp_routine", "4FB3D8", "FFFFFF", False),
        ("WFP", "ארגונים שותפים", "wfp_partners", "4FB3D8", "FFFFFF", False),
        ("WFP", "מאפיות", "wfp_bakeries", "4FB3D8", "FFFFFF", False),
        ('או"ם', 'סוכנויות או"ם מאושרות', "un_approved", "FFC000", "FFFFFF", False),
        ('או"ם', 'סוכנויות או"ם לא מוכרות (UN Sisters)', "un_unrecognized", "FF0000", "FFFFFF", False),
        (None, "בריאות", "health", "6A1BDA", "FFFFFF", True),
        (None, UNKNOWN_ENTITIES_STATUS, "unknown_entities", "C044E8", "FFFFFF", True),
        (None, "תקשורת", "communications", "21A789", "FFFFFF", True),
    ]
    if sums.get("unclassified", 0):
        row_specs.append((None, "לא מסווג", "unclassified", "808080", "FFFFFF", True))

    total = sum(sums.get(key, 0.0) for _, _, key, *_ in row_specs)

    start_row = 2
    for i, spec in enumerate(row_specs, start=start_row):
        sector, detail, key, fill_hex, font_hex, merge_detail = spec
        _write_dashboard_row(ws, i, sector, detail, sums.get(key, 0.0), total, fill_hex, font_hex, merge_detail)

    # Merge group labels vertically (NGO/WFP/UN) after writing row borders/fills.
    group_ranges = [(2, 4, "NGO", "2B8C65"), (6, 8, "WFP", "4FB3D8"), (9, 10, 'או"ם', "FFC000")]
    for r1, r2, label, fill_hex in group_ranges:
        ws.merge_cells(start_row=r1, start_column=1, end_row=r2, end_column=1)
        cell = ws.cell(r1, 1)
        cell.value = label
        cell.fill = PatternFill(fill_type="solid", fgColor=fill_hex)
        cell.font = Font(color="FFFFFF", bold=True, size=16)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        for r in range(r1, r2 + 1):
            ws.cell(r, 1).border = Border(
                left=Side(style="thin", color="000000"),
                right=Side(style="thin", color="000000"),
                top=Side(style="thin", color="000000"),
                bottom=Side(style="thin", color="000000"),
            )

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 12
    for r in range(1, ws.max_row + 1):
        ws.row_dimensions[r].height = 28

    # Keep this sheet clean: only the requested Hebrew summary table is created here.

    return sums, pd.DataFrame(source_rows)

def add_status_to_fuels_summary_workbook(fuels_summary_file, approval_index, prefixes, include_match_details=False):
    raw = BytesIO(fuels_summary_file.getvalue())
    wb = load_workbook(raw)

    if DISTRIBUTION_SHEET_NAME not in wb.sheetnames:
        raise RuntimeError(f'Could not find sheet "{DISTRIBUTION_SHEET_NAME}" in {fuels_summary_file.name}.')

    ws = wb[DISTRIBUTION_SHEET_NAME]
    headers = {
        "Cluster": ["Cluster"],
        "Agency": ["AGENCY", "Agency"],
        "Description": ["DESCRIPTION", "Description"],
    }
    header_row, cols = find_header_row_and_columns(ws, headers, max_scan_rows=5)
    if not header_row:
        raise RuntimeError('Could not find Cluster / AGENCY / DESCRIPTION headers in Distribution Summary.')

    status_col, inserted = get_or_insert_status_column(ws, header_row, cols["Description"])

    detail_headers = ["Match Type", "Match Score", "Matched Cluster/Intervention", "Matched Agency", "Matched Description", "Approval Sheet Row"]
    detail_cols = {}
    if include_match_details:
        start_col = ws.max_column + 1
        for i, h in enumerate(detail_headers):
            c = start_col + i
            detail_cols[h] = c
            cell = ws.cell(header_row, c)
            cell.value = h
            copy_cell_style(ws.cell(header_row, status_col), cell)
            ws.column_dimensions[get_column_letter(c)].width = 24

    review_rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        cluster = ws.cell(r, cols["Cluster"]).value
        agency = ws.cell(r, cols["Agency"]).value
        desc = ws.cell(r, cols["Description"]).value
        if not any([cluster, agency, desc]):
            continue

        review = review_summary_row(cluster, agency, desc, approval_index, prefixes)
        status_cell = ws.cell(r, status_col)
        status_cell.value = review["Review Status"]

        # Use the row's existing style, then color only the status cell for easy review.
        copy_cell_style(ws.cell(r, cols["Description"]), status_cell)
        status_cell.fill = status_fill(review["Review Status"])
        status_cell.font = Font(bold=True, color="000000")
        status_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        if include_match_details:
            for h in detail_headers:
                c = detail_cols[h]
                ws.cell(r, c).value = review[h]
                copy_cell_style(status_cell, ws.cell(r, c))
                ws.cell(r, c).fill = PatternFill(fill_type=None)
                ws.cell(r, c).font = Font(bold=False, color="000000")

        review_rows.append({
            "Workbook Row": r,
            "Detected Org": review["Detected Org"],
            "Cluster": cluster,
            "Agency": agency,
            "Description": desc,
            "Review Status": review["Review Status"],
            "Match Type": review["Match Type"],
            "Match Score": review["Match Score"],
            "Matched Cluster/Intervention": review["Matched Cluster/Intervention"],
            "Matched Agency": review["Matched Agency"],
            "Matched Description": review["Matched Description"],
            "Approval Sheet Row": review["Approval Sheet Row"],
        })

    # Header styling for the new status column.
    header_cell = ws.cell(header_row, status_col)
    header_cell.value = STATUS_COL_NAME
    header_cell.fill = PatternFill(fill_type="solid", fgColor="000000")
    header_cell.font = Font(color="FFFFFF", bold=True)
    header_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Status Review Log removed per final request.
    # Keep Distribution Summary usable, but filter only real header columns.
    ws.freeze_panes = "A2"
    _set_autofilter_to_real_headers(ws, header_row=header_row)

    # Build the requested Hebrew dashboard sheet from row-level Unified Fuel values
    # BEFORE deleting helper columns.
    add_status_dashboard_sheet(wb, ws, header_row=header_row)

    # Remove internal helper columns from the final Distribution Summary.
    _delete_columns_by_headers(ws, header_row, ["Fuel sum", "Description Sum"])
    _set_autofilter_to_real_headers(ws, header_row=header_row)

    # Remove legacy QA sheet if present.
    if "Status Review Log" in wb.sheetnames:
        wb.remove(wb["Status Review Log"])

    # Apply final formatting/layout requests.
    apply_final_workbook_layout(wb)

    out = BytesIO()
    wb.save(out)
    out.seek(0)

    return out, pd.DataFrame(review_rows)




# ============================================================
# Persistent approval workbook helper
# ============================================================
APPROVAL_FILE_NAME = "חלוקת דלקים לארגונים מאושרים, מסורבים ומוקפאים.xlsx"
APPROVAL_STORE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), APPROVAL_FILE_NAME) if "__file__" in globals() else APPROVAL_FILE_NAME

class LocalWorkbookFile:
    def __init__(self, path):
        self.path = path
        self.name = os.path.basename(path)
        self.size = os.path.getsize(path) if os.path.exists(path) else 0

    def getvalue(self):
        with open(self.path, "rb") as f:
            return f.read()

class BytesWorkbookFile:
    def __init__(self, name, data):
        self.name = name
        self._data = data
        self.size = len(data)

    def getvalue(self):
        return self._data

def get_saved_approval_file():
    if os.path.exists(APPROVAL_STORE_PATH):
        return LocalWorkbookFile(APPROVAL_STORE_PATH)
    return None

def save_uploaded_approval_file(uploaded_file):
    with open(APPROVAL_STORE_PATH, "wb") as f:
        f.write(uploaded_file.getvalue())
    return LocalWorkbookFile(APPROVAL_STORE_PATH)


# ============================================================
# UI: Upload → combine/calculations → add Fuel Dashboard → download
# ============================================================

left, right = st.columns([2, 1])

with left:
    st.markdown('<div class="card">', unsafe_allow_html=True)

    saved_approval = get_saved_approval_file()
    if saved_approval is not None:
        st.caption(f"Using local approval workbook: {saved_approval.name}")
    else:
        st.warning(f"Approval workbook was not found next to the app file: {APPROVAL_FILE_NAME}")

    uploads = st.file_uploader(
        "Upload Total Distribution .xlsx files",
        type=["xlsx"],
        accept_multiple_files=True,
        key="distribution_uploads",
    )
    latest_day_uploads = st.file_uploader(
        "Upload latest-day UNOPS + WFP files for Fuel Dashboard",
        type=["xlsx"],
        accept_multiple_files=True,
        key="latest_day_uploads",
    )
    st.markdown(
        '<div class="small">The app reads the approval workbook directly from the code folder. The final workbook will include Distribution Summary, Sector Summary, Fuel Dashboard, and סטטוס לפי אישורים.</div>',
        unsafe_allow_html=True,
    )
    st.markdown("</div>", unsafe_allow_html=True)

with right:
    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown("**Pipeline**")
    st.markdown(
        """
1) Combine Total Distribution files  
2) Run Fuels Cleaner calculations  
3) Create Distribution Summary  
4) Create Sector Summary  
5) Add Fuel Dashboard sheet from latest-day files  
6) Add סטטוס review from local approval list  
7) Add סטטוס לפי אישורים summary table  
8) Download final workbook
"""
    )
    st.markdown("</div>", unsafe_allow_html=True)

st.markdown("### Fuel Dashboard week dates")
col1, col2 = st.columns(2)
with col1:
    start_date = st.date_input("Week start date")
with col2:
    end_date = st.date_input("Week end date")

selected_dates = []
if start_date and end_date:
    current = start_date
    while current <= end_date:
        selected_dates.append(current)
        current += timedelta(days=1)


def _reset_cached_output():
    st.session_state.pop("final_bytes", None)
    st.session_state.pop("final_name", None)
    st.session_state.pop("debug_logs", None)
    st.session_state.pop("status_review_summary", None)
    st.session_state.pop("last_upload_key", None)


if "last_upload_key" not in st.session_state:
    st.session_state.last_upload_key = None

run_btn = st.button("Run combine + calculations", type="primary", disabled=not uploads)

upload_key_parts = []
if uploads:
    upload_key_parts.extend([f"dist:{u.name}:{u.size}" for u in uploads])
if latest_day_uploads:
    upload_key_parts.extend([f"dash:{u.name}:{u.size}" for u in latest_day_uploads])
if start_date and end_date:
    upload_key_parts.append(f"dates:{start_date.isoformat()}:{end_date.isoformat()}")
if os.path.exists(APPROVAL_STORE_PATH):
    upload_key_parts.append(f"approval:{os.path.getsize(APPROVAL_STORE_PATH)}:{os.path.getmtime(APPROVAL_STORE_PATH)}")

upload_key = "|".join(upload_key_parts)
if upload_key and st.session_state.last_upload_key != upload_key:
    st.session_state.last_upload_key = upload_key
    st.session_state.pop("final_bytes", None)
    st.session_state.pop("final_name", None)
    st.session_state.pop("debug_logs", None)
    st.session_state.pop("status_review_summary", None)

if "final_bytes" in st.session_state and st.session_state.get("final_bytes"):
    st.success("Done! Download your final file below.")
    if st.session_state.get("status_review_summary"):
        with st.expander("Status review summary"):
            st.dataframe(pd.DataFrame(st.session_state.status_review_summary), use_container_width=True)
    if st.session_state.get("debug_logs"):
        with st.expander("Fuel Dashboard debug logs"):
            st.text("\n".join(st.session_state.debug_logs))
    st.download_button(
        "Download final Excel",
        data=st.session_state.final_bytes,
        file_name=st.session_state.get("final_name", "Fuels summary.xlsx"),
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
    st.stop()

if run_btn:
    progress = st.progress(0)
    status = st.empty()
    loader = st.empty()
    show_small_loader_video(loader, "dog_running.mp4", width_px=220)

    try:
        debug_logs = []
        approval_source = get_saved_approval_file()
        if approval_source is None:
            raise RuntimeError(f"Approval/status workbook was not found next to the app file: {APPROVAL_FILE_NAME}")

        status.info("Step 1/4: Combining Total Distribution files…")
        progress.progress(5)
        combined_bytes = build_combined_workbook_bytes(uploads, status=status)

        status.info("Step 2/4: Running Fuels Cleaner calculations…")
        progress.progress(25)
        final_bytes = run_calculations_on_combined_bytes(combined_bytes, progress=progress, status=status)

        if latest_day_uploads:
            if not selected_dates:
                raise RuntimeError("Please select valid week start and end dates for the Fuel Dashboard.")

            status.info("Step 3/4: Adding Fuel Dashboard sheet…")
            progress.progress(85)

            final_bytes.seek(0)
            final_wb = load_workbook(final_bytes)
            final_wb = add_fuel_dashboard_sheet(
                final_wb,
                latest_day_uploads,
                selected_dates,
                logs=debug_logs,
                sheet_name="Fuel Dashboard",
            )
            out = BytesIO()
            final_wb.save(out)
            out.seek(0)
            final_bytes = out
        else:
            status.warning("No latest-day UNOPS/WFP files uploaded, so Fuel Dashboard was not added.")

        status.info("Step 4/4: Adding approval status review and סטטוס לפי אישורים sheet…")
        progress.progress(94)
        approval_index, prefixes, approval_df = build_approval_index(approval_source)
        final_bytes.seek(0)
        reviewed_excel, review_df = add_status_to_fuels_summary_workbook(
            BytesWorkbookFile("generated_fuels_summary.xlsx", final_bytes.getvalue()),
            approval_index,
            prefixes,
            include_match_details=False,
        )
        final_bytes = reviewed_excel

        if review_df is not None and not review_df.empty:
            status_summary_df = (
                review_df.groupby(["Detected Org", "Review Status", "Match Type"], dropna=False)
                .size()
                .reset_index(name="Count")
                .sort_values(["Detected Org", "Review Status", "Match Type"])
            )
            st.session_state.status_review_summary = status_summary_df.to_dict("records")
        else:
            st.session_state.status_review_summary = []

        progress.progress(100)
        st.session_state.final_bytes = final_bytes.getvalue()
        st.session_state.final_name = "Fuels summary.xlsx"
        st.session_state.debug_logs = debug_logs

        status.success("Done! Download your final file below.")
        if st.session_state.get("status_review_summary"):
            with st.expander("Status review summary"):
                st.dataframe(pd.DataFrame(st.session_state.status_review_summary), use_container_width=True)
        if debug_logs:
            with st.expander("Fuel Dashboard debug logs"):
                st.text("\n".join(debug_logs))
        st.download_button(
            "Download final Excel",
            data=st.session_state.final_bytes,
            file_name=st.session_state.final_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    except Exception as e:
        status.empty()
        st.exception(e)

    finally:
        loader.empty()
